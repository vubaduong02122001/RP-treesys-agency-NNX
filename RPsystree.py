# @title
import os
import glob
import time
import calendar
import shutil
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import pandas as pd
import numpy as np
from collections import defaultdict
from datetime import datetime, timedelta

import pytz
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Các import riêng cho Colab (bọc trong try để không lỗi khi chạy local)
try:
    from google.colab import drive, files
    IS_COLAB = True
except ImportError:
    IS_COLAB = False


# ============================================================
# CẤU HÌNH HỆ THỐNG
# ============================================================

@dataclass
class AppConfig:
    # MỤC TIÊU MẶC ĐỊNH
    DEFAULT_TARGET_ACTIVE_RATE = 30.0        # 30%
    DEFAULT_TARGET_NS_ACTIVE = 5_000_000.0   # 5,000,000

    """Cấu hình đường dẫn & thông số nguồn dữ liệu."""
    base_dir: str = "/content/drive/MyDrive/Database"

    # Google Sheet cấu hình
    sheet_id: str = "1ohUajfJtJvfO5D2trBStKVQkKdJUCm2ru3gW08zTrpc"
    sheet_gid: str = "2096375418"

    # Tên thư mục con cấp 1
    rawdata_folder: str = "Rawdata"
    ctv_folder: str = "Danh sách CTV"
    target_folder: str = "Mục tiêu"
    phi_t_folder: str = "Phí bảo hiểm T"
    phi_t1_folder: str = "Phí bảo hiểm T-1"

    # Tên subfolder bên trong Rawdata
    rawdata_onl_subfolder: str = "Cấp onl"
    rawdata_off_subfolder: str = "Cấp off"

    # Thư mục lưu dữ liệu giữ hạng theo Quý/Tháng
    keep_position_folder: str = "Rawdata_RP_keepposotion"

    # Tham số phân tích thời gian
    timezone: str = "Asia/Ho_Chi_Minh"

    @property
    def rawdata_dir(self) -> str:
        return os.path.join(self.base_dir, self.rawdata_folder)

    @property
    def rawdata_onl_dir(self) -> str:
        # Thư mục Rawdata/Cấp onl
        return os.path.join(self.base_dir, self.rawdata_folder, self.rawdata_onl_subfolder)

    @property
    def rawdata_off_dir(self) -> str:
        # Thư mục Rawdata/Cấp off
        return os.path.join(self.base_dir, self.rawdata_folder, self.rawdata_off_subfolder)

    @property
    def ctv_dir(self) -> str:
        return os.path.join(self.base_dir, self.ctv_folder)

    @property
    def target_dir(self) -> str:
        return os.path.join(self.base_dir, self.target_folder)

    @property
    def phi_t_dir(self) -> str:
        return os.path.join(self.base_dir, self.phi_t_folder)

    @property
    def phi_t1_dir(self) -> str:
        return os.path.join(self.base_dir, self.phi_t1_folder)

    @property
    def keep_position_dir(self) -> str:
        return os.path.join(self.base_dir, self.keep_position_folder)


# ============================================================
# UTILS
# ============================================================

def show_loading(message: str, steps: int = 25, delay: float = 0.03) -> None:
    """Thông báo trạng thái, bỏ progress-bar cho nhanh."""
    print(message)


def get_latest_excel_in_folder(folder: str,
                               required: bool = True,
                               description: str = "") -> Optional[str]:
    """
    Lấy file Excel mới nhất trong thư mục bằng cách duyệt toàn bộ.
    """
    files_list = glob.glob(os.path.join(folder, "*.xlsx")) + glob.glob(os.path.join(folder, "*.xls"))

    if not files_list:
        if required:
            raise FileNotFoundError(f"Không tìm thấy file {description} trong: {folder}")
        print(f"Không tìm thấy file {description} trong: {folder}")
        return None

    latest_file = None
    latest_mtime = -1.0

    for path in files_list:
        try:
            mtime = os.path.getmtime(path)
        except Exception as e:
            print(f"⚠ Không đọc được thời gian sửa đổi file: {path} | Lỗi: {e}")
            continue

        if mtime > latest_mtime:
            latest_mtime = mtime
            latest_file = path

    if latest_file is None:
        if required:
            raise FileNotFoundError(f"Không tìm được file hợp lệ trong: {folder}")
        print(f"Không tìm được file hợp lệ trong: {folder}")
        return None

    print(f"✓ Chọn file {description}: {os.path.basename(latest_file)}")
    return latest_file


def get_fill_by_percentage(percentage: float) -> PatternFill:
    """Trả về màu fill theo phần trăm hoàn thành."""
    if percentage < 50:
        return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # đỏ
    elif percentage < 75:
        return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # vàng
    else:
        return PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")  # xanh lá


# ============================================================
# CLASS XỬ LÝ RAWDATA + GOOGLE SHEET
# ============================================================

class InsuranceDataProcessor:
    """Xử lý dữ liệu giao dịch bảo hiểm: Rawdata + Google Sheet."""

    def __init__(self,
                 rawdata_path: str,
                 sheet_id: str,
                 gid: str,
                 channel_filter: Optional[str] = "CollaboratorApp",
                 excluded_partner: Optional[str] = None,
                 offline_folder: Optional[str] = None) -> None:
        """
        channel_filter:
            - None  : không lọc theo CHANNEL_NAME.
            - Ví dụ "CollaboratorApp": lọc theo kênh này.
        excluded_partner:
            - None  : không loại đối tác (KHÔNG lọc PVI ở bước raw).
        offline_folder:
            - Thư mục dùng để lưu file Google Sheet tải về (ví dụ Rawdata/Cấp off).
        """
        self.rawdata_path = rawdata_path
        self.sheet_id = sheet_id
        self.gid = gid
        self.channel_filter = channel_filter
        self.excluded_partner = excluded_partner
        self.offline_folder = offline_folder

        self.df_combined: Optional[pd.DataFrame] = None  # dữ liệu sau khi đã lọc PVI/PVI_VCX
        self.df_all: Optional[pd.DataFrame] = None       # dữ liệu đầy đủ, CHƯA lọc PVI/PVI_VCX

    # ---------- PUBLIC PIPELINE ----------

    def process(self) -> pd.DataFrame:
        """
        Pipeline xử lý rawdata + Google Sheet.

        - Bước 1: lọc raw theo CHANNEL_NAME (Cấp onl).
        - Bước 2: đọc Google Sheet (Cấp off) + chuẩn hóa.
        - Bước 3: kết hợp onl + off, lưu bản đầy đủ vào self.df_all.
        - Bước 4: lọc bỏ PVI/PVI_VCX trên INS_TYPE để xuất file phí T & làm báo cáo.
        """
        df_raw = self._load_rawdata()
        df_sheet = self._load_google_sheet()

        # df_full: đã kết hợp Cấp onl + Cấp off, CHƯA lọc PVI/PVI_VCX
        df_full = self._merge_data(df_raw, df_sheet)

        # Lưu bản đầy đủ (gồm cả VCX) để tính "Tổng phí gồm VCX"
        self.df_all = df_full.copy()

        # Áp dụng lọc PVI/PVI_VCX cho dữ liệu sử dụng tiếp
        df_filtered = df_full.copy()
        if "INS_TYPE" in df_filtered.columns:
            ins_upper = df_filtered["INS_TYPE"].astype(str).str.upper()
            mask = ~ins_upper.isin(["PVI", "PVI_VCX"])
            before = len(df_filtered)
            df_filtered = df_filtered[mask].copy()
            removed = before - len(df_filtered)
            print(f"✓ Đã lọc bỏ {removed:,} dòng INS_TYPE thuộc PVI/PVI_VCX (dùng cho báo cáo & file T)")
            print(f"✓ Tổng giao dịch sau lọc: {len(df_filtered):,} dòng")
            print(f"✓ Tổng phí bảo hiểm sau lọc: {df_filtered['CONTRACT_AMT'].sum():,.0f} VNĐ")
        else:
            print("⚠ Không tìm thấy cột INS_TYPE, bỏ qua bước lọc PVI/PVI_VCX.")

        self.df_combined = df_filtered
        return df_filtered

    def export_to_file(self, output_path: str) -> None:
        """Xuất df_combined (đã lọc PVI/PVI_VCX) ra file Excel Phí bảo hiểm T."""
        if self.df_combined is None:
            raise ValueError("Chưa có dữ liệu để xuất (df_combined is None).")
        show_loading("Đang xuất file Phí bảo hiểm T...")
        self.df_combined.to_excel(output_path, index=False, engine="openpyxl")
        print(f"✓ Đã xuất file phí bảo hiểm T: {output_path}")

    # ---------- INTERNAL STEPS ----------

    def _load_rawdata(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("BƯỚC 1: XỬ LÝ RAWDATA (CẤP ONL)")
        print("=" * 100)

        show_loading("Đang đọc rawdata (Cấp onl)...")
        df = pd.read_excel(self.rawdata_path, engine="openpyxl")
        print(f"✓ Tổng số dòng rawdata ban đầu (Cấp onl): {len(df):,} ")

        # Lọc theo CHANNEL_NAME nếu được cấu hình
        if self.channel_filter:
            if "CHANNEL_NAME" not in df.columns:
                raise ValueError("Rawdata thiếu cột 'CHANNEL_NAME' để lọc.")
            df = df[df["CHANNEL_NAME"] == self.channel_filter].copy()
            print(f"✓ Sau lọc CHANNEL_NAME='{self.channel_filter}': {len(df):,} dòng")
            if df.empty:
                raise ValueError("Không có dữ liệu sau khi lọc CHANNEL_NAME.")
        else:
            print("✓ Không lọc theo CHANNEL_NAME (dùng nguyên file Cấp onl).")

        # KHÔNG lọc PVI nữa ở bước raw (excluded_partner = None)
        if self.excluded_partner:
            df = self._remove_partner_records(df, self.excluded_partner)
        else:
            print("✓ Không loại đối tác nào ở bước raw (PVI vẫn giữ lại nếu có).")

        return df

    @staticmethod
    def _remove_partner_records(df: pd.DataFrame, excluded_partner: str) -> pd.DataFrame:
        """Loại bỏ các bản ghi thuộc một đối tác nhất định (ví dụ PVI)."""
        has_partner = "PARTNER_CODE" in df.columns
        has_ins_type = "INS_TYPE" in df.columns

        if not (has_partner or has_ins_type):
            print("⚠ Không tìm thấy cột PARTNER_CODE/INS_TYPE, bỏ qua lọc đối tác.")
            return df

        mask = pd.Series(True, index=df.index)
        if has_partner:
            mask &= df["PARTNER_CODE"].astype(str).str.upper().ne(excluded_partner.upper())
        if has_ins_type:
            mask &= df["INS_TYPE"].astype(str).str.upper().ne(excluded_partner.upper())

        df = df[mask].copy()
        print(f"✓ Sau loại '{excluded_partner}': {len(df):,} dòng")
        return df

    def _load_google_sheet(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("BƯỚC 2: ĐỌC GOOGLE SHEET (CẤP OFF)")
        print("=" * 100)

        url = f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/export?format=csv&gid={self.gid}"
        show_loading("Đang tải dữ liệu từ Google Sheet...")

        # Đọc toàn bộ sheet
        df_sheet_raw = pd.read_csv(url)

        # Chuẩn hóa tên cột về lower-case để map linh hoạt
        cols_lower_map = {c.strip().lower(): c for c in df_sheet_raw.columns}

        # Map logic_name -> (key_lower, internal_name)
        required_spec = [
            ("họ & tên ctv", "Họ & Tên CTV"),
            ("mã ctv", "Mã CTV"),
            ("phí bảo hiểm", "Phí Bảo Hiểm"),
            ("ctbh", "CTBH"),
            ("ngày cấp", "NGAY_CAP"),  # cột ngày cấp trên Google Sheet
        ]

        actual_cols: Dict[str, str] = {}
        missing_logical: List[str] = []

        for key_lower, internal_name in required_spec:
            if key_lower not in cols_lower_map:
                if internal_name == "NGAY_CAP":
                    missing_logical.append("Ngày cấp")
                else:
                    missing_logical.append(internal_name)
            else:
                actual_cols[internal_name] = cols_lower_map[key_lower]

        if missing_logical:
            raise ValueError(f"Google Sheet thiếu cột: {', '.join(missing_logical)}")

        # Lấy đúng các cột cần, theo tên gốc, rồi chuẩn hóa tên cột nội bộ
        df_sheet = df_sheet_raw[
            [
                actual_cols["Họ & Tên CTV"],
                actual_cols["Mã CTV"],
                actual_cols["Phí Bảo Hiểm"],
                actual_cols["CTBH"],
                actual_cols["NGAY_CAP"],
            ]
        ].copy()

        df_sheet.columns = [
            "Họ & Tên CTV",
            "Mã CTV",
            "Phí Bảo Hiểm",
            "CTBH",
            "NGAY_CAP",
        ]

        # Chuyển "ngày cấp" từ dd/mm/yyyy -> yyyymmdd (string)
        ngay_cap_dt = pd.to_datetime(
            df_sheet["NGAY_CAP"], format="%d/%m/%Y", errors="coerce"
        )
        df_sheet["NGAY_CAP"] = ngay_cap_dt.dt.strftime("%Y%m%d")
        df_sheet["NGAY_CAP"] = df_sheet["NGAY_CAP"].fillna("")

        # Xác định thư mục 'Cấp off' để lưu file Google Sheet
        if self.offline_folder is not None:
            off_dir = self.offline_folder
        else:
            raw_onl_dir = os.path.dirname(self.rawdata_path)      # .../Rawdata/Cấp onl
            raw_root_dir = os.path.dirname(raw_onl_dir)           # .../Rawdata
            off_dir = os.path.join(raw_root_dir, "Cấp off")

        os.makedirs(off_dir, exist_ok=True)

        # Đặt tên file Google Sheet tải về
        today_str = datetime.now().strftime("%Y%m%d")
        sheet_filename = f"GoogleSheet_Cap_off_{today_str}.xlsx"
        sheet_path = os.path.join(off_dir, sheet_filename)

        # Lưu bản copy Google Sheet về Rawdata/Cấp off
        df_sheet.to_excel(sheet_path, index=False, engine="openpyxl")

        print(f"✓ Đã lưu Google Sheet xuống file: {sheet_path}")
        print(f"✓ Đọc thành công {len(df_sheet):,} dòng từ Google Sheet")

        return df_sheet

    @staticmethod
    def _ensure_columns(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
        """Đảm bảo các cột đích tồn tại trong df (nếu thiếu thì tạo None)."""
        for col in cols:
            if col not in df.columns:
                df[col] = None
        return df

    def _merge_data(self, df_raw: pd.DataFrame, df_sheet: pd.DataFrame) -> pd.DataFrame:
        """
        BƯỚC 3: KẾT HỢP DỮ LIỆU RAWDATA + GOOGLE SHEET (CẤP ONL + CẤP OFF)

        - df_raw: dữ liệu online (Cấp onl) đã lọc CHANNEL_NAME.
        - df_sheet: dữ liệu offline lấy từ Google Sheet (đã có cột NGAY_CAP dạng yyyymmdd).
        - Tạo df_offline với các cột:
            + FINISH_EMPLOYEE_NM   <- 'Họ & Tên CTV'
            + FINISH_EMPLOYEE_CODE <- 'Mã CTV'
            + CONTRACT_AMT         <- 'Phí Bảo Hiểm'
            + PARTNER_CODE         <- 'CTBH'
            + INS_TYPE             <- 'CTBH'
            + DATE_WID             <- 'NGAY_CAP' (yyyymmdd)
        - Nối dọc df_raw + df_offline -> df_combined (CHƯA lọc PVI/PVI_VCX).
        - Việc lọc PVI/PVI_VCX được thực hiện ở hàm process().
        """
        print("\n" + "=" * 100)
        print("BƯỚC 3: KẾT HỢP DỮ LIỆU RAWDATA (CẤP ONL) + GOOGLE SHEET (CẤP OFF)")
        print("=" * 100)

        # Đảm bảo df_raw có đầy đủ các cột cần thiết
        target_cols = [
            "FINISH_EMPLOYEE_NM",
            "FINISH_EMPLOYEE_CODE",
            "CONTRACT_AMT",
            "PARTNER_CODE",
            "INS_TYPE",
            "DATE_WID",
        ]
        df_raw = self._ensure_columns(df_raw, target_cols)

        # Tạo khung df_offline với cùng bộ cột như df_raw
        df_offline = pd.DataFrame(columns=df_raw.columns)

        # Gán dữ liệu từ Google Sheet vào các cột tương ứng
        df_offline["FINISH_EMPLOYEE_NM"] = df_sheet["Họ & Tên CTV"]
        df_offline["FINISH_EMPLOYEE_CODE"] = df_sheet["Mã CTV"]
        df_offline["CONTRACT_AMT"] = pd.to_numeric(
            df_sheet["Phí Bảo Hiểm"], errors="coerce"
        ).fillna(0)

        # Copy - paste CTBH vào PARTNER_CODE và INS_TYPE (block offline)
        df_offline["PARTNER_CODE"] = df_sheet["CTBH"]
        df_offline["INS_TYPE"] = df_sheet["CTBH"]

        # Copy - paste NGAY_CAP (yyyymmdd) vào DATE_WID cho block offline
        if "NGAY_CAP" in df_sheet.columns:
            df_offline["DATE_WID"] = df_sheet["NGAY_CAP"].astype(str).str.strip()
        else:
            df_offline["DATE_WID"] = None

        show_loading("Đang kết hợp dữ liệu (Cấp onl + Cấp off)...")

        # Nối dọc: onl trước, off sau
        df_combined = pd.concat([df_raw, df_offline], ignore_index=True)

        print(f"✓ Số dòng online (Cấp onl): {len(df_raw):,}")
        print(f"✓ Số dòng offline từ Google Sheet (Cấp off): {len(df_offline):,}")
        print(f"✓ Tổng giao dịch trước lọc INS_TYPE: {len(df_combined):,} dòng")
        print(f"✓ Tổng phí bảo hiểm trước lọc: {df_combined['CONTRACT_AMT'].sum():,.0f} VNĐ")

        # Trả về df_combined CHƯA lọc PVI/PVI_VCX
        return df_combined


# ============================================================
# HIERARCHY + KPI / BÁO CÁO
# ============================================================

@dataclass
class ReportTargets:
    individual_targets: Dict[str, int]
    target_active_rate: float
    target_ns_active: float


@dataclass
class TimeContext:
    today: datetime
    yesterday: datetime
    week_ago: datetime
    month_ago: datetime
    days_in_month: int
    days_worked: int
    quarter_index: int
    quarter_label: str
    quarter_months: List[int]
    retention_month_labels: List[str]


# Chỉ tiêu giữ hạng theo chức danh
RETENTION_ROLE_TARGETS: Dict[str, Dict[str, int]] = {
    "Trưởng nhóm": {"premium": 100_000_000, "active": 30},
    "Trưởng phòng": {"premium": 250_000_000, "active": 70},
    "Giám đốc": {"premium": 400_000_000, "active": 100},
}


def attach_quarter_month_premiums(
    df_ctv: pd.DataFrame,
    time_ctx: TimeContext,
    config: AppConfig,
    code_column: str = "CTV_CODE_CLEAN",
) -> pd.DataFrame:
    """
    Gắn thêm các cột phí theo tháng Quý hiện tại cho từng CTV.
    Nguồn dữ liệu: /Database/Rawdata_RP_keepposotion/Qx/T{tháng}/
    """
    quarter_label = time_ctx.quarter_label
    quarter_months = time_ctx.quarter_months

    base_q_dir = os.path.join(config.keep_position_dir, quarter_label)
    os.makedirs(base_q_dir, exist_ok=True)

    print("\n" + "=" * 100)
    print(f"GẮN PHÍ THÁNG CHO BÁO CÁO GIỮ HẠNG ({quarter_label})")
    print("=" * 100)

    code_candidates = [
        "FINISH_EMPLOYEE_CODE", "EMPLOYEE_CODE", "CODE", "MA_CTV", "CTV_CODE"
    ]

    for month in quarter_months:
        month_label = f"Tháng {month}"
        month_col_name = month_label

        t_folder_candidates = [
            os.path.join(base_q_dir, f"T{month}"),
            os.path.join(base_q_dir, f"T{month:02d}"),
        ]

        month_file = None
        for folder in t_folder_candidates:
            if os.path.isdir(folder):
                try:
                    month_file = get_latest_excel_in_folder(
                        folder,
                        required=False,
                        description=f"{quarter_label}-{month_label}",
                    )
                except Exception:
                    month_file = None
                if month_file:
                    break

        if month_file is None:
            print(f"⚠ Không tìm thấy dữ liệu {month_label} trong {quarter_label}, gán 0.")
            df_ctv[month_col_name] = 0
            continue

        print(f"✓ Dùng file {os.path.basename(month_file)} cho {month_label}")
        df_month = pd.read_excel(month_file, engine="openpyxl")
        df_month.columns = df_month.columns.str.strip()
        upper = df_month.columns.str.upper()

        month_code_col = None
        for c in code_candidates:
            if c in upper.values:
                month_code_col = df_month.columns[upper.tolist().index(c)]
                break

        if not month_code_col:
            print(f"⚠ Không xác định được cột MÃ CTV trong file {month_file}, gán 0.")
            df_ctv[month_col_name] = 0
            continue

        df_month["CTV_CODE_CLEAN"] = df_month[month_code_col].astype(str).str.strip()
        if "CONTRACT_AMT" not in df_month.columns:
            print(f"⚠ File {month_file} không có cột CONTRACT_AMT, gán 0.")
            df_ctv[month_col_name] = 0
            continue

        df_month["CONTRACT_AMT"] = pd.to_numeric(
            df_month["CONTRACT_AMT"], errors="coerce"
        ).fillna(0)

        month_by_ctv = df_month.groupby("CTV_CODE_CLEAN")["CONTRACT_AMT"].sum()
        df_ctv[month_col_name] = df_ctv[code_column].map(month_by_ctv).fillna(0)
        print(f"✓ Gắn phí {month_label} cho {df_ctv[month_col_name].gt(0).sum():,} CTV")

    return df_ctv


class HierarchyBuilder:
    """Xây dựng phân cấp hệ thống CTV."""

    def __init__(self, df_ctv: pd.DataFrame) -> None:
        self.df_ctv = df_ctv.copy()
        self._build_mappings()

    def _build_mappings(self) -> None:
        df = self.df_ctv

        self.emp2level: Dict[str, str] = (
            df.set_index("EMPLOYEE_CODE")["CHANNEL_LEVEL_CODE"].astype(str).to_dict()
        )
        self.emp2ref: Dict[str, Optional[str]] = (
            df.set_index("EMPLOYEE_CODE")["REFERRAL_CODE"].to_dict()
        )
        self.emp2name: Dict[str, str] = (
            df.set_index("EMPLOYEE_CODE")["FULL_NAME"].fillna(df["USER_NAME"]).astype(str).to_dict()
        )

        grouped = df.groupby("REFERRAL_CODE")["EMPLOYEE_CODE"].apply(list)
        self.manager2subs: Dict[str, List[str]] = grouped.to_dict()

        self.level_map = {
            "LEVEL04": "Giám đốc",
            "LEVEL03": "Trưởng phòng",
            "LEVEL02": "Trưởng nhóm",
            "LEVEL01": "Cộng tác viên",
        }

        self.director_mapping: Dict[str, str] = {}
        self.director_code_to_name: Dict[str, str] = {}
        self.director_code_to_emp_code: Dict[str, str] = {}

        self.manager_code_mapping: Dict[str, str] = {}
        self.manager_code_to_name: Dict[str, str] = {}
        self.manager_to_director: Dict[str, str] = {}
        self.manager_code_to_emp_code: Dict[str, str] = {}

        self.team_lead_code_mapping: Dict[str, str] = {}
        self.team_lead_code_to_name: Dict[str, str] = {}
        self.team_lead_to_manager: Dict[str, str] = {}
        self.team_lead_code_to_emp_code: Dict[str, str] = {}

    def _find_director_emp(self, emp_code: str, max_iter: int = 20) -> Optional[str]:
        visited = set()
        current = emp_code

        for _ in range(max_iter):
            if current in visited:
                break
            visited.add(current)

            level = self.emp2level.get(current)
            if level == "LEVEL04":
                return current

            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                break
            current = ref

        return None

    def _find_manager_emp_within_2_levels(self, emp_code: str) -> Optional[str]:
        level = self.emp2level.get(emp_code)
        if level == "LEVEL03":
            return emp_code
        if level == "LEVEL04":
            return None

        current = emp_code
        for _ in range(2):
            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                return None
            ref_level = self.emp2level.get(ref)
            if ref_level == "LEVEL03":
                return ref
            if ref_level == "LEVEL04":
                return None
            current = ref
        return None

    def _find_team_lead_emp_within_2_levels(self, emp_code: str) -> Optional[str]:
        level = self.emp2level.get(emp_code)
        if level == "LEVEL02":
            return emp_code
        if level in ("LEVEL03", "LEVEL04"):
            return None

        current = emp_code
        for _ in range(2):
            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                return None
            ref_level = self.emp2level.get(ref)
            if ref_level == "LEVEL02":
                return ref
            if ref_level in ("LEVEL03", "LEVEL04"):
                return None
            current = ref
        return None

    def find_all_subordinates_2_levels(self, manager_code: str) -> set:
        subordinates = set()
        direct = self.manager2subs.get(manager_code, [])
        subordinates.update(direct)
        for sub in direct:
            indirect = self.manager2subs.get(sub, [])
            subordinates.update(indirect)
        return subordinates

    def build(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("PHÂN CẤP HỆ THỐNG CTV")
        print("=" * 100)
        show_loading("Đang xử lý phân cấp...")

        df = self.df_ctv

        self._assign_director_codes(df)
        self._assign_manager_codes(df)
        self._assign_team_lead_codes(df)

        print(
            f"✓ Đã phân cấp: "
            f"{len(self.director_mapping)} Giám đốc, "
            f"{len(self.manager_code_mapping)} Trưởng phòng, "
            f"{len(self.team_lead_code_mapping)} Trưởng nhóm"
        )

        return df

    def _assign_director_codes(self, df: pd.DataFrame) -> None:
        directors = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL04"].sort_values("EMPLOYEE_CODE").copy()

        for idx, (_, row) in enumerate(directors.iterrows(), 1):
            emp_code = row["EMPLOYEE_CODE"]
            name = row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else row["USER_NAME"]
            dir_id = f"GD{idx:02d}"
            self.director_mapping[emp_code] = dir_id
            self.director_code_to_name[dir_id] = str(name)
            self.director_code_to_emp_code[dir_id] = emp_code

        def assign_director_code_row(row):
            emp_code = row["EMPLOYEE_CODE"]
            level = row["CHANNEL_LEVEL_CODE"]
            if level == "LEVEL04":
                return self.director_mapping.get(emp_code, "GD_UNKNOWN")
            director_emp = self._find_director_emp(emp_code)
            if director_emp and director_emp in self.director_mapping:
                return self.director_mapping[director_emp]
            return "INDEPENDENT"

        df["DIRECTOR_CODE"] = df.apply(assign_director_code_row, axis=1)
        df["DIRECTOR_NAME"] = df["DIRECTOR_CODE"].map(self.director_code_to_name).fillna("")

    def _assign_manager_codes(self, df: pd.DataFrame) -> None:
        manager_counter: Dict[str, int] = defaultdict(int)

        level3_df = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL03"].copy()
        for _, row in level3_df.iterrows():
            emp_code = row["EMPLOYEE_CODE"]
            d_code = row["DIRECTOR_CODE"]
            if d_code not in ["INDEPENDENT", "GD_UNKNOWN"]:
                manager_counter[d_code] += 1
                m_code = f"{d_code}_TP{manager_counter[d_code]:02d}"
                self.manager_code_mapping[emp_code] = m_code
                self.manager_code_to_name[m_code] = (
                    row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else ""
                )
                self.manager_to_director[m_code] = d_code
                self.manager_code_to_emp_code[m_code] = emp_code

        def assign_manager_code_row(row):
            lvl = row["CHANNEL_LEVEL_CODE"]
            emp_code = row["EMPLOYEE_CODE"]
            if lvl == "LEVEL03":
                return self.manager_code_mapping.get(emp_code, "")
            if lvl == "LEVEL04":
                return ""
            manager_emp = self._find_manager_emp_within_2_levels(emp_code)
            return self.manager_code_mapping.get(manager_emp, "") if manager_emp else ""

        df["MANAGER_CODE"] = df.apply(assign_manager_code_row, axis=1)
        df["MANAGER_NAME"] = df["MANAGER_CODE"].map(self.manager_code_to_name).fillna("")

    def _assign_team_lead_codes(self, df: pd.DataFrame) -> None:
        team_lead_counter: Dict[str, int] = defaultdict(int)

        level2_df = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL02"].copy()
        for _, row in level2_df.iterrows():
            emp_code = row["EMPLOYEE_CODE"]
            m_code = row["MANAGER_CODE"]
            if m_code:
                team_lead_counter[m_code] += 1
                tl_code = f"{m_code}_TN{team_lead_counter[m_code]:02d}"
                self.team_lead_code_mapping[emp_code] = tl_code
                self.team_lead_code_to_name[tl_code] = (
                    row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else ""
                )
                self.team_lead_to_manager[tl_code] = m_code
                self.team_lead_code_to_emp_code[tl_code] = emp_code

        def assign_team_lead_code_row(row):
            lvl = row["CHANNEL_LEVEL_CODE"]
            emp_code = row["EMPLOYEE_CODE"]
            if lvl == "LEVEL02":
                return self.team_lead_code_mapping.get(emp_code, "")
            if lvl in ["LEVEL03", "LEVEL04"]:
                return ""
            tl_emp = self._find_team_lead_emp_within_2_levels(emp_code)
            return self.team_lead_code_mapping.get(tl_emp, "") if tl_emp else ""

        df["TEAM_LEAD_CODE"] = df.apply(assign_team_lead_code_row, axis=1)
        df["TEAM_LEAD_NAME"] = df["TEAM_LEAD_CODE"].map(self.team_lead_code_to_name).fillna("")


class ReportCalculator:
    """Tính toán pivot, tăng trưởng, tổng quan, giữ hạng."""

    def __init__(self,
                 df_ctv: pd.DataFrame,
                 hierarchy: HierarchyBuilder,
                 has_t1_data: bool,
                 targets: ReportTargets,
                 time_ctx: TimeContext) -> None:
        self.df_ctv = df_ctv
        self.h = hierarchy
        self.has_t1_data = has_t1_data
        self.targets = targets
        self.time_ctx = time_ctx

    def _get_target_by_name(self, name: str) -> int:
        return self.targets.individual_targets.get(str(name).strip(), 0)

    # ---------- PIVOT CHI TIẾT ----------

    def compute_pivot(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("TÍNH TOÁN DỮ LIỆU PIVOT")
        print("=" * 100)
        show_loading("Đang tính toán dữ liệu báo cáo...")

        df_ctv = self.df_ctv
        director_mapping = self.h.director_mapping
        director_code_to_name = self.h.director_code_to_name
        director_code_to_emp_code = self.h.director_code_to_emp_code
        manager_to_director = self.h.manager_to_director
        manager_code_to_name = self.h.manager_code_to_name
        manager_code_to_emp_code = self.h.manager_code_to_emp_code
        team_lead_to_manager = self.h.team_lead_to_manager
        team_lead_code_to_name = self.h.team_lead_code_to_name
        team_lead_code_to_emp_code = self.h.team_lead_code_to_emp_code

        pivot_data = []
        forecast_ratio = (
            self.time_ctx.days_in_month / self.time_ctx.days_worked
            if self.time_ctx.days_worked > 0
            else 0
        )

        level_map = self.h.level_map

        for dir_code in sorted(director_mapping.values()):
            dir_name = director_code_to_name[dir_code]
            dir_emp_code = director_code_to_emp_code[dir_code]
            d_df = df_ctv[df_ctv["DIRECTOR_CODE"] == dir_code]
            if d_df.empty:
                continue

            pivot_data.append(
                self._build_pivot_row(
                    df=d_df,
                    name=dir_name,
                    emp_code=dir_emp_code,
                    role=level_map["LEVEL04"],
                    level=0,
                    forecast_ratio=forecast_ratio,
                )
            )

            tps = [k for k, v in manager_to_director.items() if v == dir_code]
            for tp_code in sorted(tps):
                tp_name = manager_code_to_name.get(tp_code, "")
                tp_emp_code = manager_code_to_emp_code.get(tp_code, "")
                tp_df = d_df[d_df["MANAGER_CODE"] == tp_code]
                if tp_df.empty:
                    continue

                pivot_data.append(
                    self._build_pivot_row(
                        df=tp_df,
                        name=tp_name,
                        emp_code=tp_emp_code,
                        role=level_map["LEVEL03"],
                        level=1,
                        forecast_ratio=forecast_ratio,
                    )
                )

                tns = [k for k, v in team_lead_to_manager.items() if v == tp_code]
                for tn_code in sorted(tns):
                    tn_name = team_lead_code_to_name.get(tn_code, "")
                    tn_emp_code = team_lead_code_to_emp_code.get(tn_code, "")

                    tn_subs = self.h.find_all_subordinates_2_levels(tn_emp_code)
                    tn_subs.add(tn_emp_code)
                    tn_df = df_ctv[df_ctv["EMPLOYEE_CODE"].isin(tn_subs)]

                    if tn_df.empty:
                        continue

                    pivot_data.append(
                        self._build_pivot_row(
                            df=tn_df,
                            name=tn_name,
                            emp_code=tn_emp_code,
                            role=level_map["LEVEL02"],
                            level=2,
                            forecast_ratio=forecast_ratio,
                        )
                    )

        print(f"✓ Đã tính toán {len(pivot_data)} bản ghi cho báo cáo pivot")

        export_columns = [
            "Tên", "Mã CTV", "Chức danh", "Tổng số CTV", "Số Active",
            "Tổng phí T-1", "Tổng phí", "Tổng phí gồm VCX", "Mục tiêu", "Dự báo",
            "Số HĐ", "Tỷ lệ Active (%)", "Mục tiêu tỷ lệ Active (%)",
            "Năng suất TB", "Năng suất/Active", "Mục tiêu năng suất/Active",
        ]

        df_pivot = pd.DataFrame(pivot_data)
        if not df_pivot.empty:
            df_pivot = df_pivot[export_columns + ["Level"]]
        else:
            df_pivot = pd.DataFrame(columns=export_columns + ["Level"])
        return df_pivot

    def _build_pivot_row(self,
                         df: pd.DataFrame,
                         name: str,
                         emp_code: str,
                         role: str,
                         level: int,
                         forecast_ratio: float) -> Dict:
        has_t1 = self.has_t1_data
        target_active_rate = self.targets.target_active_rate
        target_ns_active = self.targets.target_ns_active

        total = len(df)
        active = int(df["IS_ACTIVE"].sum())
        premium_sum = df["TOTAL_PREMIUM"].sum()
        if "TOTAL_PREMIUM_ALL" in df.columns:
            premium_sum_all = df["TOTAL_PREMIUM_ALL"].sum()
        else:
            premium_sum_all = premium_sum

        premium_t1 = df["TOTAL_PREMIUM_T1"].sum() if has_t1 else 0
        forecast = int(premium_sum * forecast_ratio) if forecast_ratio > 0 else 0
        contracts = int(df["CONTRACT_COUNT"].sum())

        active_rate = round(active / total * 100, 2) if total > 0 else 0
        ns_avg = round(premium_sum / total, 0) if total > 0 else 0
        ns_active = round(premium_sum / active, 0) if active > 0 else 0

        target_value = self._get_target_by_name(name)

        return {
            "Tên": name,
            "Mã CTV": emp_code,
            "Chức danh": role,
            "Tổng số CTV": total,
            "Số Active": active,
            "Tổng phí T-1": int(premium_t1),
            "Tổng phí": int(premium_sum),
            "Tổng phí gồm VCX": int(premium_sum_all),
            "Mục tiêu": int(target_value),
            "Dự báo": forecast,
            "Số HĐ": contracts,
            "Tỷ lệ Active (%)": active_rate,
            "Mục tiêu tỷ lệ Active (%)": target_active_rate,
            "Năng suất TB": int(ns_avg),
            "Năng suất/Active": int(ns_active),
            "Mục tiêu năng suất/Active": int(target_ns_active),
            "Level": level,
        }

    # ---------- TĂNG TRƯỞNG TOÀN HỆ THỐNG ----------

    def compute_growth(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("PHÂN TÍCH TĂNG TRƯỞNG TOÀN HỆ THỐNG")
        print("=" * 100)

        df = self.df_ctv
        tc = self.time_ctx

        growth_data = []
        for level_code, level_name in [
            ("LEVEL04", "Giám đốc"),
            ("LEVEL03", "Trưởng phòng"),
            ("LEVEL02", "Trưởng nhóm"),
            ("LEVEL01", "Cộng tác viên"),
        ]:
            today_count = len(df[(df["CHANNEL_LEVEL_CODE"] == level_code) &
                                 (df["CREATED_DATE"] <= tc.today)])
            yesterday_count = len(df[(df["CHANNEL_LEVEL_CODE"] == level_code) &
                                     (df["CREATED_DATE"] <= tc.yesterday)])
            week_count = len(df[(df["CHANNEL_LEVEL_CODE"] == level_code) &
                                (df["CREATED_DATE"] <= tc.week_ago)])
            month_count = len(df[(df["CHANNEL_LEVEL_CODE"] == level_code) &
                                 (df["CREATED_DATE"] <= tc.month_ago)])

            d1_num = today_count - yesterday_count
            d1_pct = round(d1_num / yesterday_count * 100, 1) if yesterday_count > 0 else 0.0

            w1_num = today_count - week_count
            w1_pct = round(w1_num / week_count * 100, 1) if week_count > 0 else 0.0

            t1_num = today_count - month_count
            t1_pct = round(t1_num / month_count * 100, 2) if month_count > 0 else 0.0

            total_ctv = len(df[df["CREATED_DATE"] <= tc.today])
            ratio = round(today_count / total_ctv * 100, 2) if total_ctv > 0 else 0.0

            growth_data.append({
                "Chức danh": level_name,
                "Số lượng": today_count,
                "Tăng trưởng so với D-1": f"{d1_num} ({d1_pct}%)",
                "Tăng trưởng so với W-1": f"{w1_num} ({w1_pct}%)",
                "Tăng trưởng so với T-1": f"{t1_num} ({t1_pct}%)",
                "Tỉ trọng/Tổng CTV": f"{ratio}%",
            })

        df_growth = pd.DataFrame(growth_data)
        print(f"✓ Đã phân tích tăng trưởng cho {len(growth_data)} cấp bậc")
        return df_growth

    # ---------- CƠ CẤU THEO GIÁM ĐỐC ----------

    def compute_director_growth(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("PHÂN TÍCH CƠ CẤU THEO GIÁM ĐỐC")
        print("=" * 100)

        df = self.df_ctv
        director_mapping = self.h.director_mapping
        director_code_to_name = self.h.director_code_to_name
        level_map = self.h.level_map

        director_growth_data = []
        for dir_code in sorted(director_mapping.values()):
            dir_name = director_code_to_name[dir_code]
            d_df = df[df["DIRECTOR_CODE"] == dir_code]
            if d_df.empty:
                continue

            for level_code, level_name in level_map.items():
                count = len(d_df[d_df["CHANNEL_LEVEL_CODE"] == level_code])
                total_in_system = len(d_df)
                ratio_system = round(count / total_in_system * 100, 2) if total_in_system > 0 else 0.0
                ratio_company = round(count / len(df) * 100, 2) if len(df) > 0 else 0.0

                director_growth_data.append({
                    "Giám đốc": dir_name,
                    "Chức danh": level_name,
                    "Số lượng": count,
                    "Tỉ trọng/Hệ thống": f"{ratio_system}%",
                    "Tỉ trọng/Tổng công ty": f"{ratio_company}%",
                })

        df_director_growth = pd.DataFrame(director_growth_data)
        print(f"✓ Đã phân tích {len(director_growth_data)} bản ghi theo giám đốc")
        return df_director_growth

    # ---------- BÁO CÁO GIỮ HẠNG ----------

    def compute_retention_report(self) -> pd.DataFrame:
        """
        Tạo DataFrame dùng cho sheet 'Báo cáo giữ hạng'.

        - Mỗi dòng là 1 Giám đốc / Trưởng phòng / Trưởng nhóm.
        - Tổng phí bảo hiểm = tổng phí của 3 tháng trong Quý hiện tại
          (tức là tổng các cột Tháng ...).
        """
        print("\n" + "=" * 100)
        print("TÍNH BÁO CÁO GIỮ HẠNG")
        print("=" * 100)

        df_ctv = self.df_ctv
        h = self.h
        rows: List[Dict] = []
        month_cols = self.time_ctx.retention_month_labels

        # 1. Hệ thống có Giám đốc
        for dir_code in sorted(h.director_mapping.values()):
            dir_name = h.director_code_to_name[dir_code]
            dir_emp_code = h.director_code_to_emp_code[dir_code]

            d_df = df_ctv[df_ctv["DIRECTOR_CODE"] == dir_code]
            if d_df.empty:
                continue

            # Giám đốc
            rows.append(
                self._build_retention_row(
                    df=d_df,
                    name=dir_name,
                    emp_code=dir_emp_code,
                    role=h.level_map["LEVEL04"],
                    level=0,
                    director_code=dir_code,
                    month_cols=month_cols,
                )
            )

            # Trưởng phòng
            tps = [k for k, v in h.manager_to_director.items() if v == dir_code]
            for tp_code in sorted(tps):
                tp_name = h.manager_code_to_name.get(tp_code, "")
                tp_emp_code = h.manager_code_to_emp_code.get(tp_code, "")
                tp_df = d_df[d_df["MANAGER_CODE"] == tp_code]
                if tp_df.empty:
                    continue

                rows.append(
                    self._build_retention_row(
                        df=tp_df,
                        name=tp_name,
                        emp_code=tp_emp_code,
                        role=h.level_map["LEVEL03"],
                        level=1,
                        director_code=dir_code,
                        month_cols=month_cols,
                    )
                )

                # Trưởng nhóm
                tns = [k for k, v in h.team_lead_to_manager.items() if v == tp_code]
                for tn_code in sorted(tns):
                    tn_name = h.team_lead_code_to_name.get(tn_code, "")
                    tn_emp_code = h.team_lead_code_to_emp_code.get(tn_code, "")

                    tn_subs = h.find_all_subordinates_2_levels(tn_emp_code)
                    tn_subs.add(tn_emp_code)
                    tn_df = df_ctv[df_ctv["EMPLOYEE_CODE"].isin(tn_subs)]
                    if tn_df.empty:
                        continue

                    rows.append(
                        self._build_retention_row(
                            df=tn_df,
                            name=tn_name,
                            emp_code=tn_emp_code,
                            role=h.level_map["LEVEL02"],
                            level=2,
                            director_code=dir_code,
                            month_cols=month_cols,
                        )
                    )

        # 2. Trưởng phòng / Trưởng nhóm độc lập (DIRECTOR_CODE = 'INDEPENDENT')
        indep = df_ctv[df_ctv["DIRECTOR_CODE"] == "INDEPENDENT"].copy()
        if not indep.empty:
            # Trưởng phòng độc lập
            indep_mgr = indep[indep["CHANNEL_LEVEL_CODE"] == "LEVEL03"]
            for emp_code in sorted(indep_mgr["EMPLOYEE_CODE"].unique()):
                subs = h.find_all_subordinates_2_levels(emp_code)
                subs.add(emp_code)
                sub_df = df_ctv[df_ctv["EMPLOYEE_CODE"].isin(subs)]
                if sub_df.empty:
                    continue
                name = h.emp2name.get(emp_code, str(emp_code))
                rows.append(
                    self._build_retention_row(
                        df=sub_df,
                        name=name,
                        emp_code=str(emp_code),
                        role=h.level_map["LEVEL03"],
                        level=1,
                        director_code="INDEPENDENT",
                        month_cols=month_cols,
                    )
                )

            # Trưởng nhóm độc lập
            indep_tl = indep[indep["CHANNEL_LEVEL_CODE"] == "LEVEL02"]
            for emp_code in sorted(indep_tl["EMPLOYEE_CODE"].unique()):
                subs = h.find_all_subordinates_2_levels(emp_code)
                subs.add(emp_code)
                sub_df = df_ctv[df_ctv["EMPLOYEE_CODE"].isin(subs)]
                if sub_df.empty:
                    continue
                name = h.emp2name.get(emp_code, str(emp_code))
                rows.append(
                    self._build_retention_row(
                        df=sub_df,
                        name=name,
                        emp_code=str(emp_code),
                        role=h.level_map["LEVEL02"],
                        level=2,
                        director_code="INDEPENDENT",
                        month_cols=month_cols,
                    )
                )

        if not rows:
            cols = (
                ["Họ Tên", "Mã CTV", "Chức danh"]
                + month_cols
                + [
                    "Tổng phí bảo hiểm",
                    "Mục tiêu phí bảo hiểm",
                    "SL CTV active",
                    "Mục tiêu SL CTV active",
                    "Đánh giá",
                    "Level",
                    "DIRECTOR_CODE",
                ]
            )
            return pd.DataFrame(columns=cols)

        df_ret = pd.DataFrame(rows)
        print(f"✓ Đã tạo {len(df_ret)} dòng cho 'Báo cáo giữ hạng'")
        return df_ret

    def _build_retention_row(
        self,
        df: pd.DataFrame,
        name: str,
        emp_code: str,
        role: str,
        level: int,
        director_code: str,
        month_cols: List[str],
    ) -> Dict:
        """
        Tính 1 dòng cho báo cáo giữ hạng từ một nhóm CTV.

        Tổng phí bảo hiểm = tổng phí của các tháng trong 'month_cols'
        (tức là tổng 3 tháng của Quý hiện tại).
        """
        # Tổng phí 3 tháng
        quarter_total = 0
        month_values: Dict[str, int] = {}
        for col in month_cols:
            val = df[col].sum() if col in df.columns else 0
            month_values[col] = int(val)
            quarter_total += val

        active_ctv = int(df["IS_ACTIVE"].sum())

        target_cfg = RETENTION_ROLE_TARGETS.get(role, {"premium": 0, "active": 0})
        target_premium = target_cfg.get("premium", 0)
        target_active = target_cfg.get("active", 0)

        if target_premium == 0 and target_active == 0:
            evaluation = ""
        else:
            if quarter_total >= target_premium and active_ctv >= target_active:
                evaluation = "Đạt"
            else:
                evaluation = "Chưa đạt"

        row: Dict[str, object] = {
            "Họ Tên": name,
            "Mã CTV": emp_code,
            "Chức danh": role,
            "Tổng phí bảo hiểm": int(quarter_total),
            "Mục tiêu phí bảo hiểm": int(target_premium),
            "SL CTV active": active_ctv,
            "Mục tiêu SL CTV active": int(target_active),
            "Đánh giá": evaluation,
            "Level": level,
            "DIRECTOR_CODE": director_code,
        }

        # Gắn thêm các cột tháng
        for col, val in month_values.items():
            row[col] = val

        return row

    # ---------- BÁO CÁO TỔNG QUAN (ALL PHÍ, CÓ PVI/VCX) ----------

        # ---------- BÁO CÁO TỔNG QUAN (ALL PHÍ, CÓ PVI/VCX) ----------

    def compute_overview(self, df_all: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        - Tổng quan theo toàn công ty + từng giám đốc, tính trên TOTAL_PREMIUM_ALL.
        - Tổng quan theo ngày trong tháng: tổng phí toàn hệ thống và:
            + Hệ thống từng Giám đốc
            + Nhánh TN, TP không trong hệ thống Giám đốc
            + Unknown (CTV không thuộc nhánh TN/TP & hệ thống GĐ)
        """
        print("\n" + "=" * 100)
        print("BÁO CÁO TỔNG QUAN TOÀN HỆ THỐNG")
        print("=" * 100)
        show_loading("Đang tính toán tổng quan toàn hệ thống...")

        forecast_ratio = (
            self.time_ctx.days_in_month / self.time_ctx.days_worked
            if self.time_ctx.days_worked > 0
            else 0
        )

        def build_summary_row(df: pd.DataFrame, name: str, emp_code: str, role: str) -> Dict:
            total = len(df)
            active = int(df["IS_ACTIVE"].sum())
            premium_all = df["TOTAL_PREMIUM_ALL"].sum()
            premium_t1 = df["TOTAL_PREMIUM_T1"].sum() if self.has_t1_data else 0
            forecast = int(premium_all * forecast_ratio) if forecast_ratio > 0 else 0
            contracts = int(df["CONTRACT_COUNT"].sum())

            active_rate = round(active / total * 100, 2) if total > 0 else 0
            ns_avg = round(premium_all / total, 0) if total > 0 else 0
            ns_active = round(premium_all / active, 0) if active > 0 else 0

            target_value = self._get_target_by_name(name)

            return {
                "Tên": name,
                "Mã CTV": emp_code,
                "Chức danh": role,
                "Tổng số CTV": total,
                "Số Active": active,
                "Tổng phí T-1": int(premium_t1),
                "Tổng phí": int(premium_all),          # dùng tổng phí gồm VCX
                "Tổng phí gồm VCX": int(premium_all),  # cột mirror cho rõ
                "Mục tiêu": int(target_value),
                "Dự báo": forecast,
                "Số HĐ": contracts,
                "Tỷ lệ Active (%)": active_rate,
                "Mục tiêu tỷ lệ Active (%)": self.targets.target_active_rate,
                "Năng suất TB": int(ns_avg),
                "Năng suất/Active": int(ns_active),
                "Mục tiêu năng suất/Active": int(self.targets.target_ns_active),
            }

        # 1. Tổng quan hệ thống + theo Giám đốc
        rows = []
        rows.append(build_summary_row(self.df_ctv, "Toàn hệ thống", "", "Tổng"))

        for dir_code in sorted(self.h.director_mapping.values()):
            dir_name = self.h.director_code_to_name.get(dir_code, dir_code)
            dir_emp_code = self.h.director_code_to_emp_code.get(dir_code, "")
            d_df = self.df_ctv[self.df_ctv["DIRECTOR_CODE"] == dir_code]
            if d_df.empty:
                continue
            rows.append(build_summary_row(d_df, dir_name, dir_emp_code, "Giám đốc"))

        df_summary = pd.DataFrame(rows)

        # 2. Tổng quan theo ngày trong tháng (all phí, gồm PVI/PVI_VCX)
        df_all_local = df_all.copy()

        # Chuẩn hóa mã CTV
        if "CTV_CODE_CLEAN" not in df_all_local.columns:
            df_all_local["CTV_CODE_CLEAN"] = (
                df_all_local.get("FINISH_EMPLOYEE_CODE", "")
                .astype(str)
                .str.strip()
            )

        # Map phân cấp: GĐ / TP / TN
        df_tmp = self.df_ctv[["CTV_CODE_CLEAN", "DIRECTOR_CODE", "MANAGER_CODE", "TEAM_LEAD_CODE"]].copy()
        df_tmp["CTV_CODE_CLEAN"] = df_tmp["CTV_CODE_CLEAN"].astype(str).str.strip()
        code_to_director = df_tmp.set_index("CTV_CODE_CLEAN")["DIRECTOR_CODE"].to_dict()
        code_to_manager = df_tmp.set_index("CTV_CODE_CLEAN")["MANAGER_CODE"].to_dict()
        code_to_team_lead = df_tmp.set_index("CTV_CODE_CLEAN")["TEAM_LEAD_CODE"].to_dict()

        df_all_local["DIRECTOR_CODE"] = df_all_local["CTV_CODE_CLEAN"].map(code_to_director)
        df_all_local["MANAGER_CODE"] = df_all_local["CTV_CODE_CLEAN"].map(code_to_manager)
        df_all_local["TEAM_LEAD_CODE"] = df_all_local["CTV_CODE_CLEAN"].map(code_to_team_lead)

        # Chuẩn hóa ngày
        df_all_local["DATE"] = pd.to_datetime(
            df_all_local["DATE_WID"].astype(str).str[:8],
            format="%Y%m%d",
            errors="coerce",
        )
        start_month = self.time_ctx.today.replace(day=1)
        end_date = self.time_ctx.today

        mask = df_all_local["DATE"].notna()
        mask &= (df_all_local["DATE"] >= start_month) & (df_all_local["DATE"] <= end_date)
        df_all_local = df_all_local[mask].copy()

        if df_all_local.empty:
            df_daily = pd.DataFrame(columns=["Ngày", "Tổng phí toàn hệ thống"])
            return df_summary, df_daily

        df_all_local["DATE_ONLY"] = df_all_local["DATE"].dt.normalize()

        # Ngày trong tháng
        total_by_day = (
            df_all_local.groupby("DATE_ONLY")["CONTRACT_AMT"].sum().astype(int)
        )
        dates = sorted(total_by_day.index.unique())

        daily_df = pd.DataFrame({"Ngày": dates})
        daily_df.set_index("Ngày", inplace=True)

        # Tổng phí toàn hệ thống
        daily_df["Tổng phí toàn hệ thống"] = total_by_day.reindex(dates).fillna(0).astype(int)

        # Hệ thống từng Giám đốc
        for dir_code, dir_name in sorted(self.h.director_code_to_name.items()):
            series = (
                df_all_local[df_all_local["DIRECTOR_CODE"] == dir_code]
                .groupby("DATE_ONLY")["CONTRACT_AMT"]
                .sum()
            )
            col_name = f"Hệ thống {dir_name}"
            daily_df[col_name] = series.reindex(dates).fillna(0).astype(int)

        # Nhánh TN, TP không trong hệ thống Giám đốc
        gd_codes = set(self.h.director_mapping.values())
        dir_code_series = df_all_local["DIRECTOR_CODE"].fillna("")
        manager_series = df_all_local["MANAGER_CODE"].fillna("")
        team_lead_series = df_all_local["TEAM_LEAD_CODE"].fillna("")

        mask_director_system = dir_code_series.isin(gd_codes)
        mask_has_branch = (manager_series != "") | (team_lead_series != "")
        mask_independent_branch = (~mask_director_system) & mask_has_branch

        series_indep_branch = (
            df_all_local[mask_independent_branch]
            .groupby("DATE_ONLY")["CONTRACT_AMT"]
            .sum()
        )
        daily_df["Nhánh TN, TP không trong hệ thống Giám đốc"] = (
            series_indep_branch.reindex(dates).fillna(0).astype(int)
        )

        # Unknown: không thuộc hệ thống GĐ & không thuộc nhánh TN/TP
        mask_unknown = (~mask_director_system) & (~mask_has_branch)
        series_unknown = (
            df_all_local[mask_unknown]
            .groupby("DATE_ONLY")["CONTRACT_AMT"]
            .sum()
        )
        daily_df["Unknown"] = series_unknown.reindex(dates).fillna(0).astype(int)

        daily_df = daily_df.reset_index()
        return df_summary, daily_df


        # Tổng quan hệ thống + theo Giám đốc
        rows = []
        rows.append(build_summary_row(self.df_ctv, "Toàn hệ thống", "", "Tổng"))

        for dir_code in sorted(self.h.director_mapping.values()):
            dir_name = self.h.director_code_to_name.get(dir_code, dir_code)
            dir_emp_code = self.h.director_code_to_emp_code.get(dir_code, "")
            d_df = self.df_ctv[self.df_ctv["DIRECTOR_CODE"] == dir_code]
            if d_df.empty:
                continue
            rows.append(build_summary_row(d_df, dir_name, dir_emp_code, "Giám đốc"))

        df_summary = pd.DataFrame(rows)

        # Tổng quan theo ngày trong tháng (all phí, gồm PVI/PVI_VCX)
        df_all_local = df_all.copy()
        if "CTV_CODE_CLEAN" not in df_all_local.columns:
            df_all_local["CTV_CODE_CLEAN"] = (
                df_all_local.get("FINISH_EMPLOYEE_CODE", "").astype(str).str.strip()
            )

        code_to_director = self.df_ctv.set_index("CTV_CODE_CLEAN")["DIRECTOR_CODE"].to_dict()
        df_all_local["DIRECTOR_CODE"] = df_all_local["CTV_CODE_CLEAN"].map(code_to_director)

        df_all_local["DATE"] = pd.to_datetime(
            df_all_local["DATE_WID"].astype(str).str[:8],
            format="%Y%m%d",
            errors="coerce",
        )
        start_month = self.time_ctx.today.replace(day=1)
        end_date = self.time_ctx.today
        mask = df_all_local["DATE"].notna()
        mask &= (df_all_local["DATE"] >= start_month) & (df_all_local["DATE"] <= end_date)
        df_all_local = df_all_local[mask].copy()

        if df_all_local.empty:
            df_daily = pd.DataFrame(columns=["Ngày", "Tổng phí toàn hệ thống"])
            return df_summary, df_daily

        df_all_local["DATE_ONLY"] = df_all_local["DATE"].dt.normalize()
        total_by_day = (
            df_all_local.groupby("DATE_ONLY")["CONTRACT_AMT"].sum().astype(int)
        )

        dates = sorted(total_by_day.index.unique())
        daily_df = pd.DataFrame({"Ngày": dates})
        daily_df.set_index("Ngày", inplace=True)

        daily_df["Tổng phí toàn hệ thống"] = total_by_day.reindex(dates).fillna(0).astype(int)

        for dir_code, dir_name in sorted(self.h.director_code_to_name.items()):
            series = (
                df_all_local[df_all_local["DIRECTOR_CODE"] == dir_code]
                .groupby("DATE_ONLY")["CONTRACT_AMT"]
                .sum()
            )
            col_name = f"Hệ thống {dir_name}"
            daily_df[col_name] = series.reindex(dates).fillna(0).astype(int)

        daily_df = daily_df.reset_index()
        return df_summary, daily_df


# ============================================================
# XUẤT EXCEL + ĐỊNH DẠNG
# ============================================================

class ExcelReportWriter:
    """Ghi và định dạng file báo cáo Excel."""

    def __init__(self, output_file: str) -> None:
        self.output_file = output_file

    def write(self,
              df_pivot: pd.DataFrame,
              df_growth: pd.DataFrame,
              df_director_growth: pd.DataFrame,
              df_retention: pd.DataFrame,
              df_overview_summary: pd.DataFrame,
              df_overview_daily: pd.DataFrame) -> None:
        show_loading("Đang tạo file Excel báo cáo...")

        with pd.ExcelWriter(self.output_file, engine="openpyxl") as writer:
            # Sheet 1: Báo cáo phân cấp
            cols_export = [
                "Tên", "Mã CTV", "Chức danh", "Tổng số CTV", "Số Active",
                "Tổng phí T-1", "Tổng phí", "Tổng phí gồm VCX", "Mục tiêu", "Dự báo",
                "Số HĐ", "Tỷ lệ Active (%)", "Mục tiêu tỷ lệ Active (%)",
                "Năng suất TB", "Năng suất/Active", "Mục tiêu năng suất/Active",
            ]
            df_pivot[cols_export].to_excel(writer, sheet_name="Báo cáo phân cấp", index=False)

            # Sheet 2, 3
            df_growth.to_excel(writer, sheet_name="Phân tích tăng trưởng", index=False)
            df_director_growth.to_excel(writer, sheet_name="Theo Giám đốc", index=False)

            # Sheet 4: Báo cáo giữ hạng
            month_cols = [c for c in df_retention.columns if c.startswith("Tháng ")]
            prefix_cols = ["Họ Tên", "Mã CTV", "Chức danh"]
            suffix_cols = [
                "Tổng phí bảo hiểm", "Mục tiêu phí bảo hiểm",
                "SL CTV active", "Mục tiêu SL CTV active", "Đánh giá",
            ]
            retention_cols = prefix_cols + month_cols + suffix_cols
            df_retention_export = df_retention[retention_cols].copy()
            df_retention_export.to_excel(
                writer, sheet_name="Báo cáo giữ hạng", index=False
            )

            # Sheet 5: Tổng quan
            # BỎ cột "Tổng phí gồm VCX" chỉ trên sheet Tổng quan
            if "Tổng phí gồm VCX" in df_overview_summary.columns:
                df_overview_summary_export = df_overview_summary.drop(
                    columns=["Tổng phí gồm VCX"]
                ).copy()
            else:
                df_overview_summary_export = df_overview_summary.copy()

            df_overview_summary_export.to_excel(
                writer, sheet_name="Tổng quan", index=False, startrow=0
            )

            startrow_daily = len(df_overview_summary_export) + 2
            if not df_overview_daily.empty:
                df_overview_daily.to_excel(
                    writer,
                    sheet_name="Tổng quan",
                    index=False,
                    startrow=startrow_daily
                )

        self._format_workbook(df_pivot, df_growth, df_director_growth,
                              df_retention, df_overview_summary, df_overview_daily)

    def _format_workbook(self,
                         df_pivot: pd.DataFrame,
                         df_growth: pd.DataFrame,
                         df_director_growth: pd.DataFrame,
                         df_retention: pd.DataFrame,
                         df_overview_summary: pd.DataFrame,
                         df_overview_daily: pd.DataFrame) -> None:
        print("\n" + "=" * 100)
        print("ĐỊNH DẠNG FILE EXCEL")
        print("=" * 100)

        show_loading("Đang định dạng Excel...")

        wb = load_workbook(self.output_file)
        border_white = Border(
            left=Side(style="thin", color="FFFFFF"),
            right=Side(style="thin", color="FFFFFF"),
            top=Side(style="thin", color="FFFFFF"),
            bottom=Side(style="thin", color="FFFFFF"),
        )

        header_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        gd_fill = PatternFill(start_color="31869B", end_color="31869B", fill_type="solid")
        tp_fill = PatternFill(start_color="60497A", end_color="60497A", fill_type="solid")
        tn_fill = PatternFill(start_color="B1A0C7", end_color="B1A0C7", fill_type="solid")
        green_light = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # ------------------ Sheet 1: Báo cáo phân cấp ------------------
        ws1 = wb["Báo cáo phân cấp"]

        # Header
        for cell in ws1[1]:
            cell.fill = header_blue
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_white

        pivot_records = df_pivot.to_dict("records")

        for idx, row in enumerate(pivot_records, start=2):
            level = row["Level"]
            name_cell = ws1.cell(row=idx, column=1)

            if level == 0:
                base_font = Font(bold=True, size=11, color="FFFFFF")
                level_fill = gd_fill
            elif level == 1:
                name_cell.value = "  " + str(name_cell.value)
                base_font = Font(bold=True, size=11, color="FFFFFF")
                level_fill = tp_fill
            elif level == 2:
                name_cell.value = "    " + str(name_cell.value)
                base_font = Font(size=11, color="000000")
                level_fill = tn_fill
            else:
                base_font = Font(size=11, color="000000")
                level_fill = tn_fill

            # Fill toàn dòng (trừ các cột có logic đặc biệt)
            for col in range(1, 17):
                cell = ws1.cell(row=idx, column=col)
                if col not in [7, 8, 10, 12, 15]:
                    cell.fill = level_fill
                    cell.font = base_font
                else:
                    cell.font = Font(bold=base_font.bold, size=11, color="000000")
                cell.border = border_white
                if col in [1, 2, 3]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Tổng phí (col 7), Tổng phí gồm VCX (col 8), Mục tiêu (col 9), Dự báo (col 10)
            premium_val = row["Tổng phí"]
            premium_all_val = row["Tổng phí gồm VCX"]
            target_val = row["Mục tiêu"]
            forecast_val = row["Dự báo"]

            premium_cell = ws1.cell(row=idx, column=7)
            premium_all_cell = ws1.cell(row=idx, column=8)
            target_cell = ws1.cell(row=idx, column=9)
            forecast_cell = ws1.cell(row=idx, column=10)

            if target_val > 0:
                pct = (premium_val / target_val) * 100
                pct_all = (premium_all_val / target_val) * 100
                pct_forecast = (forecast_val / target_val) * 100
            else:
                pct = 0.0
                pct_all = 0.0
                pct_forecast = 0.0

            # Tổng phí (không gồm VCX) + % so với mục tiêu
            premium_cell.value = f"{premium_val:,.0f} ({pct:.0f}%)"
            premium_cell.fill = get_fill_by_percentage(pct)
            premium_cell.font = Font(bold=base_font.bold, size=11, color="000000")

            # Tổng phí gồm VCX + % so với mục tiêu (logic giống Tổng phí)
            premium_all_cell.value = f"{premium_all_val:,.0f} ({pct_all:.0f}%)"
            premium_all_cell.fill = get_fill_by_percentage(pct_all)
            premium_all_cell.font = Font(bold=base_font.bold, size=11, color="000000")

            # Mục tiêu
            target_cell.number_format = "#,##0"
            target_cell.fill = level_fill
            target_cell.font = base_font

            # Dự báo + % so với mục tiêu
            forecast_cell.value = f"{forecast_val:,.0f} ({pct_forecast:.0f}%)"
            forecast_cell.fill = get_fill_by_percentage(pct_forecast)
            forecast_cell.font = Font(bold=base_font.bold, size=11, color="000000")

            # Tỷ lệ Active (col 12) & mục tiêu (col 13)
            active_rate = float(row["Tỷ lệ Active (%)"])
            active_target = float(row["Mục tiêu tỷ lệ Active (%)"])
            active_rate_cell = ws1.cell(row=idx, column=12)
            active_target_cell = ws1.cell(row=idx, column=13)

            if active_target > 0:
                active_pct = (active_rate / active_target) * 100
            else:
                active_pct = 0.0

            active_rate_cell.value = f"{active_rate:.2f}% ({active_pct:.0f}%)"
            active_rate_cell.fill = get_fill_by_percentage(active_pct)
            active_rate_cell.font = Font(bold=base_font.bold, size=11, color="000000")
            active_target_cell.value = f"{active_target:.2f}%"
            active_target_cell.number_format = '0.00"%"'
            active_target_cell.fill = level_fill
            active_target_cell.font = base_font

            # Năng suất/Active (col 15) & mục tiêu (col 16)
            ns_active = row["Năng suất/Active"]
            ns_target = row["Mục tiêu năng suất/Active"]
            ns_active_cell = ws1.cell(row=idx, column=15)
            ns_target_cell = ws1.cell(row=idx, column=16)

            if ns_target > 0:
                ns_pct = (ns_active / ns_target) * 100
            else:
                ns_pct = 0.0

            ns_active_cell.value = f"{ns_active:,.0f} ({ns_pct:.0f}%)"
            ns_active_cell.fill = get_fill_by_percentage(ns_pct)
            ns_active_cell.font = Font(bold=base_font.bold, size=11, color="000000")
            ns_target_cell.number_format = "#,##0"
            ns_target_cell.fill = level_fill
            ns_target_cell.font = base_font

            # Format số cho các cột tiền / số HĐ / NS TB
            for col in [6, 14]:
                ws1.cell(row=idx, column=col).number_format = "#,##0"

        # Set width + freeze panes
        ws1.column_dimensions["A"].width = 30
        ws1.column_dimensions["B"].width = 20
        ws1.column_dimensions["C"].width = 15
        ws1.column_dimensions["D"].width = 13
        ws1.column_dimensions["E"].width = 11
        ws1.column_dimensions["F"].width = 15
        ws1.column_dimensions["G"].width = 22
        ws1.column_dimensions["H"].width = 22
        ws1.column_dimensions["I"].width = 18
        ws1.column_dimensions["J"].width = 15
        ws1.column_dimensions["K"].width = 10
        ws1.column_dimensions["L"].width = 22
        ws1.column_dimensions["M"].width = 22
        ws1.column_dimensions["N"].width = 14
        ws1.column_dimensions["O"].width = 22
        ws1.column_dimensions["P"].width = 24
        ws1.freeze_panes = "A2"

        # ------------------ Sheet 2: Phân tích tăng trưởng ------------------
        ws2 = wb["Phân tích tăng trưởng"]

        for cell in ws2[1]:
            cell.fill = header_blue
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_white

        growth_records = df_growth.to_dict("records")
        for row_idx, row in enumerate(growth_records, start=2):
            chuc_danh = row["Chức danh"]
            fill_color = green_light if chuc_danh == "Giám đốc" else white_fill
            font_style = Font(bold=True, size=11) if chuc_danh == "Giám đốc" else Font(size=11)

            for col_idx in range(1, 7):
                cell = ws2.cell(row=row_idx, column=col_idx)
                cell.fill = fill_color
                cell.font = font_style
                cell.border = border_white
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 1 else "center",
                    vertical="center"
                )

        ws2.column_dimensions["A"].width = 15
        ws2.column_dimensions["B"].width = 12
        ws2.column_dimensions["C"].width = 22
        ws2.column_dimensions["D"].width = 22
        ws2.column_dimensions["E"].width = 22
        ws2.column_dimensions["F"].width = 18
        ws2.freeze_panes = "A2"

        # ------------------ Sheet 3: Theo Giám đốc ------------------
        ws3 = wb["Theo Giám đốc"]
        for cell in ws3[1]:
            cell.fill = header_blue
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_white

        director_growth_records = df_director_growth.to_dict("records")
        for row_idx, _ in enumerate(director_growth_records, start=2):
            for col_idx in range(1, 6):
                cell = ws3.cell(row=row_idx, column=col_idx)
                cell.border = border_white
                cell.font = Font(size=11)
                cell.alignment = Alignment(
                    horizontal="center" if col_idx > 2 else "left",
                    vertical="center"
                )

        for col in ["A", "B", "C", "D", "E"]:
            ws3.column_dimensions[col].width = 20
        ws3.freeze_panes = "A2"

        # ------------------ Sheet 4: Báo cáo giữ hạng ------------------
        ws4 = wb["Báo cáo giữ hạng"]

        header_names = [cell.value for cell in ws4[1]]
        month_cols = [name for name in header_names if isinstance(name, str) and name.startswith("Tháng ")]

        # Xác định index các cột cần so sánh mục tiêu
        premium_col_name = "Tổng phí bảo hiểm"
        premium_target_col_name = "Mục tiêu phí bảo hiểm"
        active_col_name = "SL CTV active"
        active_target_col_name = "Mục tiêu SL CTV active"

        premium_col_idx = header_names.index(premium_col_name) + 1 if premium_col_name in header_names else None
        premium_target_col_idx = header_names.index(premium_target_col_name) + 1 if premium_target_col_name in header_names else None
        active_col_idx = header_names.index(active_col_name) + 1 if active_col_name in header_names else None
        active_target_col_idx = header_names.index(active_target_col_name) + 1 if active_target_col_name in header_names else None

        for cell in ws4[1]:
            cell.fill = header_blue
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)
            cell.border = border_white

        retention_records = df_retention.to_dict("records")

        for row_idx, row in enumerate(retention_records, start=2):
            level = row.get("Level", 2)
            name_cell = ws4.cell(row=row_idx, column=1)

            if level == 0:
                base_font = Font(bold=True, size=11, color="FFFFFF")
                level_fill = gd_fill
            elif level == 1:
                name_cell.value = " " + str(name_cell.value)
                base_font = Font(bold=True, size=11, color="FFFFFF")
                level_fill = tp_fill
            else:
                name_cell.value = " " + str(name_cell.value)
                base_font = Font(size=11, color="000000")
                level_fill = tn_fill

            for col_idx, col_name in enumerate(header_names, start=1):
                cell = ws4.cell(row=row_idx, column=col_idx)
                cell.fill = level_fill

                if col_name in ["Tổng phí bảo hiểm", "Mục tiêu phí bảo hiểm",
                                "SL CTV active", "Mục tiêu SL CTV active"] + month_cols:
                    cell.font = Font(bold=base_font.bold, size=11, color="000000")
                    cell.number_format = "#,##0"
                else:
                    cell.font = base_font

                cell.border = border_white
                if col_name in ["Họ Tên", "Mã CTV", "Chức danh"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Logic so sánh & tô màu cho Tổng phí bảo hiểm
            if premium_col_idx and premium_target_col_idx:
                premium_val = row.get(premium_col_name, 0) or 0
                premium_target_val = row.get(premium_target_col_name, 0) or 0
                if premium_target_val > 0:
                    pct_premium = (premium_val / premium_target_val) * 100
                else:
                    pct_premium = 0.0

                premium_cell = ws4.cell(row=row_idx, column=premium_col_idx)
                premium_cell.value = f"{premium_val:,.0f} ({pct_premium:.0f}%)"
                premium_cell.fill = get_fill_by_percentage(pct_premium)
                premium_cell.font = Font(bold=True, size=11, color="000000")

            # Logic so sánh & tô màu cho SL CTV active
            if active_col_idx and active_target_col_idx:
                active_val = row.get(active_col_name, 0) or 0
                active_target_val = row.get(active_target_col_name, 0) or 0
                if active_target_val > 0:
                    pct_active = (active_val / active_target_val) * 100
                else:
                    pct_active = 0.0

                active_cell = ws4.cell(row=row_idx, column=active_col_idx)
                active_cell.value = f"{active_val:,.0f} ({pct_active:.0f}%)"
                active_cell.fill = get_fill_by_percentage(pct_active)
                active_cell.font = Font(bold=True, size=11, color="000000")

        # Set width + freeze panes cho sheet giữ hạng
        ws4.column_dimensions["A"].width = 26
        ws4.column_dimensions["B"].width = 18
        ws4.column_dimensions["C"].width = 15
        ws4.column_dimensions["D"].width = 12
        ws4.column_dimensions["E"].width = 12
        ws4.column_dimensions["F"].width = 12
        ws4.column_dimensions["G"].width = 20
        ws4.column_dimensions["H"].width = 20
        ws4.column_dimensions["I"].width = 16
        ws4.column_dimensions["J"].width = 22
        ws4.column_dimensions["K"].width = 18
        ws4.freeze_panes = "A2"

        # Bảng CHỈ TIÊU bên phải (TN / TP / GD)
        target_start_row = 2
        target_start_col = len(header_names) + 2  # lệch một cột trống cho dễ nhìn

        target_headers = ["Chức danh", "Mục tiêu phí bảo hiểm", "SL CTV active", "SL TN"]
        for j, title in enumerate(target_headers):
            cell = ws4.cell(row=target_start_row, column=target_start_col + j, value=title)
            cell.fill = header_blue
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_white

        role_rows = [
            ("Trưởng nhóm", "TN"),
            ("Trưởng phòng", "TP"),
            ("Giám đốc", "GD"),
        ]
        for i, (role_name, short_name) in enumerate(role_rows, start=1):
            cfg = RETENTION_ROLE_TARGETS.get(role_name, {"premium": 0, "active": 0})
            r = target_start_row + i

            c_role = ws4.cell(row=r, column=target_start_col, value=short_name)
            c_role.border = border_white
            c_role.alignment = Alignment(horizontal="center", vertical="center")

            c_premium = ws4.cell(row=r, column=target_start_col + 1,
                                 value=cfg.get("premium", 0))
            c_premium.number_format = "#,##0"
            c_premium.border = border_white
            c_premium.alignment = Alignment(horizontal="right", vertical="center")

            c_active = ws4.cell(row=r, column=target_start_col + 2,
                                value=cfg.get("active", 0))
            c_active.number_format = "#,##0"
            c_active.border = border_white
            c_active.alignment = Alignment(horizontal="right", vertical="center")

            sl_tn_value = 100 if role_name == "Giám đốc" else None
            c_tn = ws4.cell(row=r, column=target_start_col + 3, value=sl_tn_value)
            if sl_tn_value is not None:
                c_tn.number_format = "#,##0"
            c_tn.border = border_white
            c_tn.alignment = Alignment(horizontal="right", vertical="center")

        note_row = target_start_row + len(role_rows) + 2
        note_cell = ws4.cell(row=note_row, column=target_start_col,
                             value="Phí bảo hiểm không gồm VCX")
        note_cell.font = Font(italic=True, size=10)

        # ------------------ Sheet 5: Tổng quan ------------------
        ws5 = wb["Tổng quan"]

        # Border đen mảnh cho riêng sheet Tổng quan (line border)
        border_thin = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        summary_rows = len(df_overview_summary)
        summary_last_row = 0

        # ===== Phần bảng summary (trên cùng) =====
        if summary_rows > 0:
            summary_header = [cell.value for cell in ws5[1]]
            n_cols_summary = len(summary_header)

            # Header
            for col_idx in range(1, n_cols_summary + 1):
                cell = ws5.cell(row=1, column=col_idx)
                cell.fill = header_blue
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center",
                                           wrap_text=True)
                cell.border = border_thin

            summary_last_row = 1 + summary_rows

            # Body: căn & border
            for row_idx in range(2, summary_last_row + 1):
                for col_idx in range(1, n_cols_summary + 1):
                    cell = ws5.cell(row=row_idx, column=col_idx)
                    cell.border = border_thin
                    if col_idx in [1, 2, 3]:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.font = Font(size=11)

            # Định dạng số
            name_to_idx = {
                name: idx + 1
                for idx, name in enumerate(summary_header)
                if isinstance(name, str)
            }
            numeric_fields = {
                "Tổng số CTV",
                "Số Active",
                "Tổng phí T-1",
                "Mục tiêu",
                "Dự báo",
                "Số HĐ",
                "Năng suất TB",
                "Năng suất/Active",
                "Mục tiêu năng suất/Active",
            }
            perc_fields = {"Tỷ lệ Active (%)", "Mục tiêu tỷ lệ Active (%)"}

            for row_idx in range(2, summary_last_row + 1):
                for col_name, col_idx in name_to_idx.items():
                    cell = ws5.cell(row=row_idx, column=col_idx)
                    if col_name in numeric_fields:
                        cell.number_format = "#,##0"
                    elif col_name in perc_fields:
                        cell.number_format = '0.00"%"'

            # Logic so sánh & màu sắc: Tổng phí, Tỷ lệ Active, Năng suất/Active
            if "Tổng phí" in name_to_idx:
                tong_phi_idx = name_to_idx["Tổng phí"]
            else:
                tong_phi_idx = None

            active_idx = name_to_idx.get("Tỷ lệ Active (%)")
            active_target_idx = name_to_idx.get("Mục tiêu tỷ lệ Active (%)")
            ns_idx = name_to_idx.get("Năng suất/Active")
            ns_target_idx = name_to_idx.get("Mục tiêu năng suất/Active")

            overview_records = df_overview_summary.to_dict("records")
            for row_offset, row_data in enumerate(overview_records, start=2):
                # Tổng phí vs Mục tiêu
                if tong_phi_idx is not None:
                    premium_val = row_data.get("Tổng phí", 0) or 0
                    target_val = row_data.get("Mục tiêu", 0) or 0
                    pct = (premium_val / target_val * 100) if target_val > 0 else 0.0

                    premium_cell = ws5.cell(row=row_offset, column=tong_phi_idx)
                    premium_cell.value = f"{premium_val:,.0f} ({pct:.0f}%)"
                    premium_cell.fill = get_fill_by_percentage(pct)
                    premium_cell.font = Font(bold=True, size=11, color="000000")
                    premium_cell.alignment = Alignment(horizontal="right", vertical="center")
                    premium_cell.border = border_thin

                # Tỷ lệ Active (%) vs Mục tiêu tỷ lệ Active (%)
                if active_idx is not None and active_target_idx is not None:
                    active_rate = float(row_data.get("Tỷ lệ Active (%)", 0) or 0)
                    active_target = float(row_data.get("Mục tiêu tỷ lệ Active (%)", 0) or 0)
                    active_pct = (active_rate / active_target * 100) if active_target > 0 else 0.0

                    active_cell = ws5.cell(row=row_offset, column=active_idx)
                    active_cell.value = f"{active_rate:.2f}% ({active_pct:.0f}%)"
                    active_cell.fill = get_fill_by_percentage(active_pct)
                    active_cell.font = Font(bold=True, size=11, color="000000")
                    active_cell.alignment = Alignment(horizontal="right", vertical="center")
                    active_cell.border = border_thin

                    target_cell = ws5.cell(row=row_offset, column=active_target_idx)
                    target_cell.value = f"{active_target:.2f}%"
                    target_cell.number_format = '0.00"%"'
                    target_cell.alignment = Alignment(horizontal="right", vertical="center")
                    target_cell.border = border_thin

                # Năng suất/Active vs Mục tiêu năng suất/Active
                if ns_idx is not None and ns_target_idx is not None:
                    ns_val = int(row_data.get("Năng suất/Active", 0) or 0)
                    ns_target = int(row_data.get("Mục tiêu năng suất/Active", 0) or 0)
                    ns_pct = (ns_val / ns_target * 100) if ns_target > 0 else 0.0

                    ns_cell = ws5.cell(row=row_offset, column=ns_idx)
                    ns_cell.value = f"{ns_val:,.0f} ({ns_pct:.0f}%)"
                    ns_cell.fill = get_fill_by_percentage(ns_pct)
                    ns_cell.font = Font(bold=True, size=11, color="000000")
                    ns_cell.alignment = Alignment(horizontal="right", vertical="center")
                    ns_cell.border = border_thin

                    ns_target_cell = ws5.cell(row=row_offset, column=ns_target_idx)
                    ns_target_cell.number_format = "#,##0"
                    ns_target_cell.alignment = Alignment(horizontal="right", vertical="center")
                    ns_target_cell.border = border_thin

        # ===== Phần bảng theo ngày (daily) =====
        if not df_overview_daily.empty:
            daily_header_row = summary_last_row + 2 if summary_last_row > 0 else 1
            daily_data_start = daily_header_row + 1
            daily_data_end = daily_data_start + len(df_overview_daily) - 1
            n_daily_cols = df_overview_daily.shape[1]

            # Header daily
            for col_idx in range(1, n_daily_cols + 1):
                cell = ws5.cell(row=daily_header_row, column=col_idx)
                cell.fill = header_blue
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border_thin

            # Data daily
            for row_idx in range(daily_data_start, daily_data_end + 1):
                for col_idx in range(1, n_daily_cols + 1):
                    cell = ws5.cell(row=row_idx, column=col_idx)
                    cell.font = Font(size=11)
                    cell.border = border_thin
                    if col_idx == 1:
                        # Ngày: hiển thị dd/mm/yyyy
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.number_format = "dd/mm/yyyy"
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"

        # ===== Auto-fit cột (fit gọn, giới hạn độ rộng) =====
        max_col = ws5.max_column
        max_row = ws5.max_row
        for col_idx in range(1, max_col + 1):
            col_letter = ws5.cell(row=1, column=col_idx).column_letter
            max_length = 0
            for row_idx in range(1, max_row + 1):
                cell = ws5.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell_str = str(cell.value)
                    if len(cell_str) > max_length:
                        max_length = len(cell_str)
            # Giới hạn cho gọn, không quá rộng
            adjusted_width = min(max_length + 2, 35)
            ws5.column_dimensions[col_letter].width = adjusted_width

        # Freeze panes
        ws5.freeze_panes = "A2"

        wb.save(self.output_file)
        print("✓ Đã định dạng xong file Excel báo cáo")


# ============================================================
# MAIN PIPELINE (ORCHESTRATOR)
# ============================================================

def build_time_context(config: AppConfig) -> Tuple[TimeContext, datetime, datetime]:
    tz = pytz.timezone(config.timezone)
    now_vn = datetime.now(tz)

    print("=" * 100)
    print("BÁO CÁO PHÂN CẤP BẢO HIỂM - PHIÊN BẢN TỐI ƯU")
    print("=" * 100)
    print(f"\nThời gian hệ thống: {now_vn.strftime('%d/%m/%Y %H:%M:%S')}")

    current_day = now_vn.day
    current_month = now_vn.month
    current_year = now_vn.year
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    days_worked = current_day - 1

    print(f"Số ngày làm việc (mặc định): {days_worked}/{days_in_month} ngày")

    confirm = input(
        f"\nXác nhận ngày báo cáo {current_day}/{current_month}/{current_year}? (y/n): "
    ).strip().lower()
    if confirm != "y":
        current_day = int(input("Nhập ngày: "))
        current_month = int(input("Nhập tháng: "))
        current_year = int(input("Nhập năm: "))
        days_in_month = calendar.monthrange(current_year, current_month)[1]
        days_worked = current_day - 1

    today_report = datetime(current_year, current_month, current_day)
    report_date_for_premium = today_report - timedelta(days=1)

    print(f"\n✓ Ngày báo cáo: {today_report.strftime('%d/%m/%Y')}")
    print(f"✓ Ngày dùng cho tên file (D-1): {report_date_for_premium.strftime('%d/%m/%Y')}")
    print(f"✓ Số ngày làm việc: {days_worked}/{days_in_month} ngày")

    today = today_report
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Xác định Quý dựa trên tháng của ngày D-1
    month_for_quarter = report_date_for_premium.month
    quarter_index = (month_for_quarter - 1) // 3 + 1
    quarter_label = f"Q{quarter_index}"
    quarter_months = [3 * (quarter_index - 1) + i for i in range(1, 4)]
    retention_month_labels = [f"Tháng {m}" for m in quarter_months]

    print(f"✓ Thuộc quý: {quarter_label} (các tháng: {', '.join(str(m) for m in quarter_months)})")

    tc = TimeContext(
        today=today,
        yesterday=yesterday,
        week_ago=week_ago,
        month_ago=month_ago,
        days_in_month=days_in_month,
        days_worked=days_worked,
        quarter_index=quarter_index,
        quarter_label=quarter_label,
        quarter_months=quarter_months,
        retention_month_labels=retention_month_labels,
    )
    return tc, today_report, report_date_for_premium


def main():
    if IS_COLAB:
        drive.mount("/content/drive")
    show_loading("Đang khởi tạo hệ thống...")

    config = AppConfig()
    os.makedirs(config.phi_t_dir, exist_ok=True)
    os.makedirs(config.rawdata_onl_dir, exist_ok=True)
    os.makedirs(config.rawdata_off_dir, exist_ok=True)
    os.makedirs(config.keep_position_dir, exist_ok=True)

    time_ctx, today_report, report_date_for_premium = build_time_context(config)

    # --------------------------------------------------------
    # 1. Đọc rawdata onl & Google Sheet, merge thành file phí T
    # --------------------------------------------------------
    raw_file = get_latest_excel_in_folder(
        config.rawdata_onl_dir,
        required=True,
        description="rawdata online (Cấp onl)"
    )

    processor = InsuranceDataProcessor(
        rawdata_path=raw_file,
        sheet_id=config.sheet_id,
        gid=config.sheet_gid,
        channel_filter="CollaboratorApp",
        excluded_partner=None,
        offline_folder=config.rawdata_off_dir,
    )

    # df_premium: đã lọc PVI/PVI_VCX (dùng cho báo cáo chính)
    df_premium = processor.process()
    # df_premium_all: CHƯA lọc PVI/PVI_VCX (dùng để tính "Tổng phí gồm VCX" và Tổng quan)
    df_premium_all = processor.df_all.copy() if processor.df_all is not None else df_premium.copy()

    premium_output_filename = (
        f"Phí bảo hiểm {report_date_for_premium.day:02d}_"
        f"{report_date_for_premium.month:02d}_"
        f"{report_date_for_premium.year}.xlsx"
    )
    premium_output_path = os.path.join(config.phi_t_dir, premium_output_filename)
    processor.export_to_file(premium_output_path)

    # Copy thêm vào thư mục giữ hạng /Rawdata_RP_keepposotion/Qx/T{tháng}
    quarter_label = time_ctx.quarter_label
    month_for_quarter = report_date_for_premium.month
    keep_q_dir = os.path.join(config.keep_position_dir, quarter_label)
    keep_t_dir = os.path.join(keep_q_dir, f"T{month_for_quarter}")
    os.makedirs(keep_q_dir, exist_ok=True)
    os.makedirs(keep_t_dir, exist_ok=True)
    keep_t_path = os.path.join(keep_t_dir, premium_output_filename)
    try:
        shutil.copy2(premium_output_path, keep_t_path)
        print(f"✓ Đã sao chép file phí T vào thư mục giữ hạng: {keep_t_path}")
    except Exception as e:
        print(f"⚠ Không sao chép được file phí T vào thư mục giữ hạng: {e}")

    print(f"\n✓ ĐÃ TẠO FILE PHÍ BẢO HIỂM T: {premium_output_filename}")

    # --------------------------------------------------------
    # 2. Đọc CTV, T-1, mục tiêu
    # --------------------------------------------------------
    print("\n" + "=" * 100)
    print("ĐỌC CÁC FILE BỔ SUNG TỪ DATABASE")
    print("=" * 100)

    ctv_filename = get_latest_excel_in_folder(
        config.ctv_dir, required=True, description="danh sách CTV"
    )
    df_ctv = pd.read_excel(ctv_filename, engine="openpyxl")
    df_ctv["CREATED_DATE"] = pd.to_datetime(df_ctv["CREATED_DATE"])

    premium_t1_filename = get_latest_excel_in_folder(
        config.phi_t1_dir, required=False, description="phí T-1"
    )
    has_t1_data = premium_t1_filename is not None
    df_premium_t1 = None
    if has_t1_data:
        df_premium_t1 = pd.read_excel(premium_t1_filename, engine="openpyxl")
        print(f"✓ Đọc file T-1: {len(df_premium_t1):,} giao dịch")
    else:
        print("⚠ Không tìm thấy file phí T-1")

    target_filename = get_latest_excel_in_folder(
        config.target_dir, required=True, description="mục tiêu"
    )
    df_target = pd.read_excel(target_filename, engine="openpyxl")

    target_dict: Dict[str, int] = {}
    for _, row in df_target.iterrows():
        name = str(row["Fullname"]).strip()
        target_value = row["Mục tiêu"]
        target_dict[name] = int(target_value) if pd.notna(target_value) else 0

    print(f"✓ Đã load {len(target_dict):,} mục tiêu cá nhân")

    # --------------------------------------------------------
    # 3. Nhập mục tiêu chung (có mặc định 30% và 5,000,000)
    # --------------------------------------------------------
    print("\n" + "=" * 100)
    print("NHẬP MỤC TIÊU CHUNG")
    print("=" * 100)
    print(
        f"\nNhấn Enter để dùng mặc định: "
        f"Tỷ lệ Active = {config.DEFAULT_TARGET_ACTIVE_RATE}%, "
        f"Năng suất/Active = {config.DEFAULT_TARGET_NS_ACTIVE:,.0f} VNĐ"
    )

    target_active_rate_input = input("\nNhập MỤC TIÊU TỶ LỆ ACTIVE (VD: 30 cho 30%): ").strip()
    if target_active_rate_input:
        target_active_rate = float(target_active_rate_input)
    else:
        target_active_rate = config.DEFAULT_TARGET_ACTIVE_RATE

    target_ns_active_input = input("Nhập MỤC TIÊU NĂNG SUẤT/ACTIVE (VD: 5000000): ").strip()
    if target_ns_active_input:
        target_ns_active = float(target_ns_active_input.replace(",", ""))
    else:
        target_ns_active = config.DEFAULT_TARGET_NS_ACTIVE

    print(f"\n✓ Mục tiêu tỷ lệ Active: {target_active_rate}%")
    print(f"✓ Mục tiêu năng suất/Active: {target_ns_active:,.0f}")

    # --------------------------------------------------------
    # 4. Chuẩn hóa dữ liệu phí & merge vào CTV
    # --------------------------------------------------------
    print("\n" + "=" * 100)
    print("XỬ LÝ DỮ LIỆU CHO BÁO CÁO PHÂN CẤP")
    print("=" * 100)

    # Chuẩn hóa df_premium (đã lọc PVI/PVI_VCX)
    df_premium.columns = df_premium.columns.str.strip()
    df_premium_col_upper = df_premium.columns.str.upper()

    code_candidates = [
        "FINISH_EMPLOYEE_CODE", "EMPLOYEE_CODE", "CODE", "MA_CTV", "CTV_CODE"
    ]
    premium_code_col = None
    for c in code_candidates:
        if c in df_premium_col_upper.values:
            premium_code_col = df_premium.columns[df_premium_col_upper.tolist().index(c)]
            break

    if not premium_code_col:
        premium_code_col = input("Nhập cột MÃ CTV trong df_premium: ").strip()

    # Chuẩn hóa df_premium_all (trước khi lọc PVI/PVI_VCX) để tính "Tổng phí gồm VCX"
    df_premium_all.columns = df_premium_all.columns.str.strip()
    df_premium_all["CTV_CODE_CLEAN"] = df_premium_all[premium_code_col].astype(str).str.strip()

    # Sau đó mới gán CTV_CODE_CLEAN cho df_premium (đã lọc)
    df_premium["CTV_CODE_CLEAN"] = df_premium[premium_code_col].astype(str).str.strip()

    # T-1
    if has_t1_data:
        df_premium_t1.columns = df_premium_t1.columns.str.strip()
        df_premium_t1_col_upper = df_premium_t1.columns.str.upper()

        premium_t1_code_col = None
        for c in code_candidates:
            if c in df_premium_t1_col_upper.values:
                premium_t1_code_col = df_premium_t1.columns[
                    df_premium_t1_col_upper.tolist().index(c)
                ]
                break

        if not premium_t1_code_col:
            premium_t1_code_col = premium_code_col

        df_premium_t1["CTV_CODE_CLEAN"] = (
            df_premium_t1[premium_t1_code_col].astype(str).str.strip()
        )
        premium_t1_by_ctv = (
            df_premium_t1.groupby("CTV_CODE_CLEAN")["CONTRACT_AMT"].sum().reset_index()
        )
        premium_t1_by_ctv.columns = ["CTV_CODE_CLEAN", "TOTAL_PREMIUM_T1"]

    # Merge vào df_ctv
    df_ctv.columns = df_ctv.columns.str.strip()
    df_ctv["CTV_CODE_CLEAN"] = df_ctv["EMPLOYEE_CODE"].astype(str).str.strip()

    # Active chỉ tính trên dữ liệu đã lọc (không tính PVI/VCX)
    active_ctv_codes = df_premium["CTV_CODE_CLEAN"].unique()
    df_ctv["IS_ACTIVE"] = df_ctv["CTV_CODE_CLEAN"].isin(active_ctv_codes)

    # Tổng phí (KHÔNG gồm VCX, sau khi lọc PVI/PVI_VCX)
    premium_by_ctv = (
        df_premium.groupby("CTV_CODE_CLEAN")["CONTRACT_AMT"]
        .agg(["sum", "count"])
        .reset_index()
    )
    premium_by_ctv.columns = ["CTV_CODE_CLEAN", "TOTAL_PREMIUM", "CONTRACT_COUNT"]

    # Tổng phí gồm VCX (tính trên df_premium_all CHƯA lọc PVI/PVI_VCX)
    premium_all_by_ctv = (
        df_premium_all.groupby("CTV_CODE_CLEAN")["CONTRACT_AMT"]
        .sum()
        .reset_index()
    )
    premium_all_by_ctv.columns = ["CTV_CODE_CLEAN", "TOTAL_PREMIUM_ALL"]

    df_ctv = df_ctv.merge(premium_by_ctv, on="CTV_CODE_CLEAN", how="left")
    df_ctv["TOTAL_PREMIUM"] = df_ctv["TOTAL_PREMIUM"].fillna(0)
    df_ctv["CONTRACT_COUNT"] = df_ctv["CONTRACT_COUNT"].fillna(0)

    df_ctv = df_ctv.merge(premium_all_by_ctv, on="CTV_CODE_CLEAN", how="left")
    df_ctv["TOTAL_PREMIUM_ALL"] = df_ctv["TOTAL_PREMIUM_ALL"].fillna(0)

    if has_t1_data:
        df_ctv = df_ctv.merge(premium_t1_by_ctv, on="CTV_CODE_CLEAN", how="left")
        df_ctv["TOTAL_PREMIUM_T1"] = df_ctv["TOTAL_PREMIUM_T1"].fillna(0)
    else:
        df_ctv["TOTAL_PREMIUM_T1"] = 0

    # Gắn thêm phí theo từng tháng trong Quý hiện tại cho báo cáo giữ hạng
    df_ctv = attach_quarter_month_premiums(df_ctv, time_ctx, config)

    # --------------------------------------------------------
    # 5. Phân cấp & tính toán báo cáo
    # --------------------------------------------------------
    hierarchy = HierarchyBuilder(df_ctv)
    df_ctv = hierarchy.build()

    targets = ReportTargets(
        individual_targets=target_dict,
        target_active_rate=target_active_rate,
        target_ns_active=target_ns_active,
    )
    calculator = ReportCalculator(
        df_ctv=df_ctv,
        hierarchy=hierarchy,
        has_t1_data=has_t1_data,
        targets=targets,
        time_ctx=time_ctx,
    )

    df_pivot = calculator.compute_pivot()
    df_growth = calculator.compute_growth()
    df_director_growth = calculator.compute_director_growth()
    df_retention = calculator.compute_retention_report()
    df_overview_summary, df_overview_daily = calculator.compute_overview(df_premium_all)

    # --------------------------------------------------------
    # 6. Ghi Excel
    # --------------------------------------------------------
    output_file = os.path.join(
        config.base_dir,
        f"Bao_Cao_{today_report.day:02d}_{today_report.month:02d}_{today_report.year}.xlsx",
    )
    writer = ExcelReportWriter(output_file)
    writer.write(df_pivot, df_growth, df_director_growth,
                 df_retention, df_overview_summary, df_overview_daily)

    print("\n" + "=" * 100)
    print("HOÀN THÀNH")
    print("=" * 100)

    print(f"\n✓ File phí bảo hiểm T: {premium_output_filename}")
    print(f"✓ File báo cáo phân cấp: {os.path.basename(output_file)}")
    print(f"✓ Tổng số bản ghi pivot: {len(df_pivot):,}")
    print(
        f"✓ Số người có mục tiêu: "
        f"{len([x for x in df_pivot.to_dict('records') if x['Mục tiêu'] > 0])}"
    )
    print(f"✓ Số dòng Báo cáo giữ hạng: {len(df_retention):,}")
    print(f"✓ Mục tiêu tỷ lệ Active: {target_active_rate}%")
    print(f"✓ Mục tiêu năng suất/Active: {target_ns_active:,.0f}")

    if IS_COLAB:
        print("\nĐang tải file về máy...")
        files.download(output_file)

    print("\nHOÀN THÀNH TẤT CẢ!")

    return (df_premium, df_pivot, df_growth, df_director_growth,
            df_retention, df_overview_summary, df_overview_daily)


if __name__ == "__main__":
    (df_premium, df_pivot, df_growth, df_director_growth,
     df_retention, df_overview_summary, df_overview_daily) = main()
