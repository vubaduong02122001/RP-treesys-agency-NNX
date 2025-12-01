"""Optimized insurance reporting pipeline."""
from __future__ import annotations

import calendar
import os
import shutil
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
import pytz
from openpyxl.styles import PatternFill

try:
    from google.colab import drive, files  # type: ignore

    IS_COLAB = True
except ImportError:  # pragma: no cover - environment specific
    IS_COLAB = False


@dataclass
class AppConfig:
    """Application configuration values."""

    DEFAULT_TARGET_ACTIVE_RATE: float = 30.0
    DEFAULT_TARGET_NS_ACTIVE: float = 5_000_000.0
    base_dir: str = "/content/drive/MyDrive/Database"
    sheet_id: str = "1ohUajfJtJvfO5D2trBStKVQkKdJUCm2ru3gW08zTrpc"
    sheet_gid: str = "2096375418"
    rawdata_folder: str = "Rawdata"
    ctv_folder: str = "Danh sách CTV"
    target_folder: str = "Mục tiêu"
    phi_t_folder: str = "Phí bảo hiểm T"
    phi_t1_folder: str = "Phí bảo hiểm T-1"
    rawdata_onl_subfolder: str = "Cấp onl"
    rawdata_off_subfolder: str = "Cấp off"
    keep_position_folder: str = "Rawdata_RP_keepposotion"
    timezone: str = "Asia/Ho_Chi_Minh"

    @property
    def rawdata_dir(self) -> str:
        return os.path.join(self.base_dir, self.rawdata_folder)

    @property
    def rawdata_onl_dir(self) -> str:
        return os.path.join(self.base_dir, self.rawdata_folder, self.rawdata_onl_subfolder)

    @property
    def rawdata_off_dir(self) -> str:
        return os.path.join(self.base_dir, self.rawdata_folder, self.rawdata_off_subfolder)

    @property
    def ctv_dir(self) -> str:
        return os.path.join(self.base_dir, self.ctv_folder)

    @property
    def target_dir(self) -> str:
        return os.path.join(self.base_dir, self.target_folder)

    @property
    def phi_t_dir(self) -> str:
        return os.path.join(self.base_dir, self.phi_t_folder)

    @property
    def phi_t1_dir(self) -> str:
        return os.path.join(self.base_dir, self.phi_t1_folder)

    @property
    def keep_position_dir(self) -> str:
        return os.path.join(self.base_dir, self.keep_position_folder)


@dataclass
class ReportTargets:
    individual_targets: Dict[str, int]
    target_active_rate: float
    target_ns_active: float


@dataclass
class TimeContext:
    today: datetime
    yesterday: datetime
    week_ago: datetime
    month_ago: datetime
    days_in_month: int
    days_worked: int
    quarter_index: int
    quarter_label: str
    quarter_months: List[int]
    retention_month_labels: List[str]


RETENTION_ROLE_TARGETS: Dict[str, Dict[str, int]] = {
    "Trưởng nhóm": {"premium": 100_000_000, "active": 30},
    "Trưởng phòng": {"premium": 250_000_000, "active": 70},
    "Giám đốc": {"premium": 400_000_000, "active": 100},
}


def show_loading(message: str, *_: object, **__: object) -> None:
    print(message)


def _ensure_directories(paths: Iterable[str | Path]) -> None:
    for path in paths:
        Path(path).mkdir(parents=True, exist_ok=True)


def _latest_excel(folder: str, *, required: bool, description: str) -> Optional[str]:
    folder_path = Path(folder)
    files = list(folder_path.glob("*.xlsx")) + list(folder_path.glob("*.xls"))
    if not files:
        if required:
            raise FileNotFoundError(f"Không tìm thấy file {description} trong: {folder}")
        print(f"Không tìm thấy file {description} trong: {folder}")
        return None

    newest = max(files, key=lambda f: f.stat().st_mtime, default=None)
    if newest is None:
        if required:
            raise FileNotFoundError(f"Không tìm thấy file {description} trong: {folder}")
        return None

    print(f"✓ Chọn file {description}: {newest.name}")
    return str(newest)


def _case_insensitive_lookup(columns: Sequence[str], keys: Sequence[str]) -> Dict[str, str]:
    lower_map = {col.strip().lower(): col for col in columns}
    return {key: lower_map[key.lower()] for key in keys if key.lower() in lower_map}


def get_fill_by_percentage(percentage: float) -> PatternFill:
    if percentage < 50:
        color = "FF0000"
    elif percentage < 75:
        color = "FFFF00"
    else:
        color = "00B050"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


class InsuranceDataProcessor:
    """Load and merge online raw data with Google Sheet offline data."""

    def __init__(
        self,
        rawdata_path: str,
        sheet_id: str,
        gid: str,
        channel_filter: Optional[str] = "CollaboratorApp",
        excluded_partner: Optional[str] = None,
        offline_folder: Optional[str] = None,
    ) -> None:
        self.rawdata_path = rawdata_path
        self.sheet_id = sheet_id
        self.gid = gid
        self.channel_filter = channel_filter
        self.excluded_partner = excluded_partner
        self.offline_folder = offline_folder

        self.df_combined: Optional[pd.DataFrame] = None
        self.df_all: Optional[pd.DataFrame] = None

    def process(self) -> pd.DataFrame:
        df_raw = self._load_rawdata()
        df_sheet = self._load_google_sheet()

        df_full = self._merge_data(df_raw, df_sheet)
        self.df_all = df_full.copy()

        if "INS_TYPE" in df_full.columns:
            ins_upper = df_full["INS_TYPE"].astype(str).str.upper()
            filtered = df_full.loc[~ins_upper.isin(["PVI", "PVI_VCX"])]
            removed = len(df_full) - len(filtered)
            print(
                f"✓ Đã lọc bỏ {removed:,} dòng INS_TYPE thuộc PVI/PVI_VCX (dùng cho báo cáo & file T)"
            )
            print(f"✓ Tổng giao dịch sau lọc: {len(filtered):,} dòng")
            print(f"✓ Tổng phí bảo hiểm sau lọc: {filtered['CONTRACT_AMT'].sum():,.0f} VNĐ")
            self.df_combined = filtered
        else:
            print("⚠ Không tìm thấy cột INS_TYPE, bỏ qua bước lọc PVI/PVI_VCX.")
            self.df_combined = df_full

        return self.df_combined

    def export_to_file(self, output_path: str) -> None:
        if self.df_combined is None:
            raise ValueError("Chưa có dữ liệu để xuất (df_combined is None).")
        show_loading("Đang xuất file Phí bảo hiểm T...")
        self.df_combined.to_excel(output_path, index=False, engine="openpyxl")
        print(f"✓ Đã xuất file phí bảo hiểm T: {output_path}")

    def _load_rawdata(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("BƯỚC 1: XỬ LÝ RAWDATA (CẤP ONL)")
        print("=" * 100)

        show_loading("Đang đọc rawdata (Cấp onl)...")
        df = pd.read_excel(self.rawdata_path, engine="openpyxl")
        print(f"✓ Tổng số dòng rawdata ban đầu (Cấp onl): {len(df):,} ")

        if self.channel_filter:
            if "CHANNEL_NAME" not in df.columns:
                raise ValueError("Rawdata thiếu cột 'CHANNEL_NAME' để lọc.")
            df = df.loc[df["CHANNEL_NAME"] == self.channel_filter]
            print(f"✓ Sau lọc CHANNEL_NAME='{self.channel_filter}': {len(df):,} dòng")
            if df.empty:
                raise ValueError("Không có dữ liệu sau khi lọc CHANNEL_NAME.")
        else:
            print("✓ Không lọc theo CHANNEL_NAME (dùng nguyên file Cấp onl).")

        if self.excluded_partner:
            df = self._remove_partner_records(df, self.excluded_partner)
        else:
            print("✓ Không loại đối tác nào ở bước raw (PVI vẫn giữ lại nếu có).")
        return df

    @staticmethod
    def _remove_partner_records(df: pd.DataFrame, excluded_partner: str) -> pd.DataFrame:
        partner_col = "PARTNER_CODE" if "PARTNER_CODE" in df.columns else None
        ins_col = "INS_TYPE" if "INS_TYPE" in df.columns else None
        if not partner_col and not ins_col:
            print("⚠ Không tìm thấy cột PARTNER_CODE/INS_TYPE, bỏ qua lọc đối tác.")
            return df

        excluded_upper = excluded_partner.upper()
        mask = pd.Series(True, index=df.index)
        if partner_col:
            mask &= df[partner_col].astype(str).str.upper().ne(excluded_upper)
        if ins_col:
            mask &= df[ins_col].astype(str).str.upper().ne(excluded_upper)

        filtered = df.loc[mask]
        print(f"✓ Sau loại '{excluded_partner}': {len(filtered):,} dòng")
        return filtered

    def _load_google_sheet(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("BƯỚC 2: ĐỌC GOOGLE SHEET (CẤP OFF)")
        print("=" * 100)

        url = f"https://docs.google.com/spreadsheets/d/{self.sheet_id}/export?format=csv&gid={self.gid}"
        show_loading("Đang tải dữ liệu từ Google Sheet...")
        df_sheet_raw = pd.read_csv(url)

        required_keys = ["họ & tên ctv", "mã ctv", "phí bảo hiểm", "ctbh", "ngày cấp"]
        mapping = _case_insensitive_lookup(df_sheet_raw.columns, required_keys)
        missing = [k for k in required_keys if k not in mapping]
        if missing:
            raise ValueError(f"Google Sheet thiếu cột: {', '.join(missing)}")

        df_sheet = df_sheet_raw[[mapping[key] for key in required_keys]].copy()
        df_sheet.columns = [
            "Họ & Tên CTV",
            "Mã CTV",
            "Phí Bảo Hiểm",
            "CTBH",
            "NGAY_CAP",
        ]

        ngay_cap = pd.to_datetime(df_sheet["NGAY_CAP"], format="%d/%m/%Y", errors="coerce")
        df_sheet["NGAY_CAP"] = ngay_cap.dt.strftime("%Y%m%d").fillna("")

        offline_dir = Path(self.offline_folder or Path(self.rawdata_path).parent.parent / "Cấp off")
        offline_dir.mkdir(parents=True, exist_ok=True)

        sheet_path = offline_dir / f"GoogleSheet_Cap_off_{datetime.now().strftime('%Y%m%d')}.xlsx"
        df_sheet.to_excel(sheet_path, index=False, engine="openpyxl")

        print(f"✓ Đã lưu Google Sheet xuống file: {sheet_path}")
        print(f"✓ Đọc thành công {len(df_sheet):,} dòng từ Google Sheet")
        return df_sheet

    @staticmethod
    def _ensure_columns(df: pd.DataFrame, cols: Sequence[str]) -> pd.DataFrame:
        missing = [col for col in cols if col not in df.columns]
        if missing:
            df = df.copy()
            for col in missing:
                df[col] = None
        return df

    def _merge_data(self, df_raw: pd.DataFrame, df_sheet: pd.DataFrame) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("BƯỚC 3: KẾT HỢP DỮ LIỆU RAWDATA (CẤP ONL) + GOOGLE SHEET (CẤP OFF)")
        print("=" * 100)

        target_cols = [
            "FINISH_EMPLOYEE_NM",
            "FINISH_EMPLOYEE_CODE",
            "CONTRACT_AMT",
            "PARTNER_CODE",
            "INS_TYPE",
            "DATE_WID",
        ]
        df_raw = self._ensure_columns(df_raw, target_cols)

        df_offline = pd.DataFrame({col: None for col in df_raw.columns})
        df_offline["FINISH_EMPLOYEE_NM"] = df_sheet["Họ & Tên CTV"]
        df_offline["FINISH_EMPLOYEE_CODE"] = df_sheet["Mã CTV"]
        df_offline["CONTRACT_AMT"] = pd.to_numeric(df_sheet["Phí Bảo Hiểm"], errors="coerce").fillna(0)
        df_offline["PARTNER_CODE"] = df_sheet["CTBH"]
        df_offline["INS_TYPE"] = df_sheet["CTBH"]
        df_offline["DATE_WID"] = df_sheet.get("NGAY_CAP", "").astype(str).str.strip()

        show_loading("Đang kết hợp dữ liệu (Cấp onl + Cấp off)...")
        df_combined = pd.concat([df_raw, df_offline], ignore_index=True)

        print(f"✓ Số dòng online (Cấp onl): {len(df_raw):,}")
        print(f"✓ Số dòng offline từ Google Sheet (Cấp off): {len(df_offline):,}")
        print(f"✓ Tổng giao dịch trước lọc INS_TYPE: {len(df_combined):,} dòng")
        print(f"✓ Tổng phí bảo hiểm trước lọc: {df_combined['CONTRACT_AMT'].sum():,.0f} VNĐ")
        return df_combined


def attach_quarter_month_premiums(
    df_ctv: pd.DataFrame,
    time_ctx: TimeContext,
    config: AppConfig,
    code_column: str = "CTV_CODE_CLEAN",
) -> pd.DataFrame:
    quarter_label = time_ctx.quarter_label
    quarter_months = time_ctx.quarter_months

    base_q_dir = Path(config.keep_position_dir) / quarter_label
    base_q_dir.mkdir(parents=True, exist_ok=True)

    print("\n" + "=" * 100)
    print(f"GẮN PHÍ THÁNG CHO BÁO CÁO GIỮ HẠNG ({quarter_label})")
    print("=" * 100)

    code_candidates = [
        "FINISH_EMPLOYEE_CODE",
        "EMPLOYEE_CODE",
        "CODE",
        "MA_CTV",
        "CTV_CODE",
    ]

    for month in quarter_months:
        month_label = f"Tháng {month}"
        t_folder_candidates = [base_q_dir / f"T{month}", base_q_dir / f"T{month:02d}"]
        month_file: Optional[str] = None
        for folder in t_folder_candidates:
            if folder.is_dir():
                found = _latest_excel(str(folder), required=False, description=f"{quarter_label}-{month_label}")
                if found:
                    month_file = found
                    break

        if month_file is None:
            print(f"⚠ Không tìm thấy dữ liệu {month_label} trong {quarter_label}, gán 0.")
            df_ctv[month_label] = 0
            continue

        print(f"✓ Dùng file {Path(month_file).name} cho {month_label}")
        df_month = pd.read_excel(month_file, engine="openpyxl")
        df_month.columns = df_month.columns.str.strip()
        upper = df_month.columns.str.upper()

        month_code_col = None
        for candidate in code_candidates:
            if candidate in upper.values:
                month_code_col = df_month.columns[upper.tolist().index(candidate)]
                break

        if not month_code_col:
            print(f"⚠ Không xác định được cột MÃ CTV trong file {month_file}, gán 0.")
            df_ctv[month_label] = 0
            continue

        df_month["CTV_CODE_CLEAN"] = df_month[month_code_col].astype(str).str.strip()
        if "CONTRACT_AMT" not in df_month.columns:
            print(f"⚠ File {month_file} không có cột CONTRACT_AMT, gán 0.")
            df_ctv[month_label] = 0
            continue

        df_month["CONTRACT_AMT"] = pd.to_numeric(df_month["CONTRACT_AMT"], errors="coerce").fillna(0)
        month_by_ctv = df_month.groupby("CTV_CODE_CLEAN")["CONTRACT_AMT"].sum()
        df_ctv[month_label] = df_ctv[code_column].map(month_by_ctv).fillna(0)
        print(f"✓ Gắn phí {month_label} cho {df_ctv[month_label].gt(0).sum():,} CTV")

    return df_ctv


class HierarchyBuilder:
    def __init__(self, df_ctv: pd.DataFrame) -> None:
        self.df_ctv = df_ctv.copy()
        self._build_mappings()

    def _build_mappings(self) -> None:
        df = self.df_ctv
        self.emp2level: Dict[str, str] = df.set_index("EMPLOYEE_CODE")["CHANNEL_LEVEL_CODE"].astype(str).to_dict()
        self.emp2ref: Dict[str, Optional[str]] = df.set_index("EMPLOYEE_CODE")["REFERRAL_CODE"].to_dict()
        self.emp2name: Dict[str, str] = (
            df.set_index("EMPLOYEE_CODE")["FULL_NAME"].fillna(df["USER_NAME"]).astype(str).to_dict()
        )

        grouped = df.groupby("REFERRAL_CODE")["EMPLOYEE_CODE"].apply(list)
        self.manager2subs: Dict[str, List[str]] = grouped.to_dict()

        self.level_map = {
            "LEVEL04": "Giám đốc",
            "LEVEL03": "Trưởng phòng",
            "LEVEL02": "Trưởng nhóm",
            "LEVEL01": "Cộng tác viên",
        }

        self.director_mapping: Dict[str, str] = {}
        self.director_code_to_name: Dict[str, str] = {}
        self.director_code_to_emp_code: Dict[str, str] = {}

        self.manager_code_mapping: Dict[str, str] = {}
        self.manager_code_to_name: Dict[str, str] = {}
        self.manager_to_director: Dict[str, str] = {}
        self.manager_code_to_emp_code: Dict[str, str] = {}

        self.team_lead_code_mapping: Dict[str, str] = {}
        self.team_lead_code_to_name: Dict[str, str] = {}
        self.team_lead_to_manager: Dict[str, str] = {}
        self.team_lead_code_to_emp_code: Dict[str, str] = {}

    def _find_director_emp(self, emp_code: str, max_iter: int = 20) -> Optional[str]:
        visited = set()
        current = emp_code
        for _ in range(max_iter):
            if current in visited:
                break
            visited.add(current)
            level = self.emp2level.get(current)
            if level == "LEVEL04":
                return current
            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                break
            current = ref
        return None

    def _find_manager_emp_within_2_levels(self, emp_code: str) -> Optional[str]:
        level = self.emp2level.get(emp_code)
        if level == "LEVEL03":
            return emp_code
        if level == "LEVEL04":
            return None

        current = emp_code
        for _ in range(2):
            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                return None
            ref_level = self.emp2level.get(ref)
            if ref_level == "LEVEL03":
                return ref
            if ref_level == "LEVEL04":
                return None
            current = ref
        return None

    def _find_team_lead_emp_within_2_levels(self, emp_code: str) -> Optional[str]:
        level = self.emp2level.get(emp_code)
        if level == "LEVEL02":
            return emp_code
        if level in ("LEVEL03", "LEVEL04"):
            return None

        current = emp_code
        for _ in range(2):
            ref = self.emp2ref.get(current)
            if ref is None or (isinstance(ref, float) and np.isnan(ref)):
                return None
            ref_level = self.emp2level.get(ref)
            if ref_level == "LEVEL02":
                return ref
            if ref_level in ("LEVEL03", "LEVEL04"):
                return None
            current = ref
        return None

    def find_all_subordinates_2_levels(self, manager_code: str) -> set:
        subordinates = set()
        direct = self.manager2subs.get(manager_code, [])
        subordinates.update(direct)
        for sub in direct:
            subordinates.update(self.manager2subs.get(sub, []))
        return subordinates

    def build(self) -> pd.DataFrame:
        print("\n" + "=" * 100)
        print("PHÂN CẤP HỆ THỐNG CTV")
        print("=" * 100)
        show_loading("Đang xử lý phân cấp...")

        df = self.df_ctv
        self._assign_director_codes(df)
        self._assign_manager_codes(df)
        self._assign_team_lead_codes(df)

        print(
            f"✓ Đã phân cấp: "
            f"{len(self.director_mapping)} Giám đốc, "
            f"{len(self.manager_code_mapping)} Trưởng phòng, "
            f"{len(self.team_lead_code_mapping)} Trưởng nhóm"
        )
        return df

    def _assign_director_codes(self, df: pd.DataFrame) -> None:
        directors = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL04"].sort_values("EMPLOYEE_CODE")

        for idx, (_, row) in enumerate(directors.iterrows(), 1):
            emp_code = row["EMPLOYEE_CODE"]
            name = row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else row["USER_NAME"]
            dir_id = f"GD{idx:02d}"
            self.director_mapping[emp_code] = dir_id
            self.director_code_to_name[dir_id] = str(name)
            self.director_code_to_emp_code[dir_id] = emp_code

        def assign_director_code_row(row: pd.Series) -> str:
            emp_code = row["EMPLOYEE_CODE"]
            level = row["CHANNEL_LEVEL_CODE"]
            if level == "LEVEL04":
                return self.director_mapping.get(emp_code, "GD_UNKNOWN")
            director_emp = self._find_director_emp(emp_code)
            if director_emp and director_emp in self.director_mapping:
                return self.director_mapping[director_emp]
            return "INDEPENDENT"

        df["DIRECTOR_CODE"] = df.apply(assign_director_code_row, axis=1)
        df["DIRECTOR_NAME"] = df["DIRECTOR_CODE"].map(self.director_code_to_name).fillna("")

    def _assign_manager_codes(self, df: pd.DataFrame) -> None:
        manager_counter: Dict[str, int] = {}

        level3_df = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL03"]
        for _, row in level3_df.iterrows():
            emp_code = row["EMPLOYEE_CODE"]
            d_code = row["DIRECTOR_CODE"]
            if d_code not in ["INDEPENDENT", "GD_UNKNOWN"]:
                manager_counter[d_code] = manager_counter.get(d_code, 0) + 1
                m_code = f"{d_code}_TP{manager_counter[d_code]:02d}"
                self.manager_code_mapping[emp_code] = m_code
                self.manager_code_to_name[m_code] = row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else ""
                self.manager_to_director[m_code] = d_code
                self.manager_code_to_emp_code[m_code] = emp_code

        def assign_manager_code_row(row: pd.Series) -> str:
            lvl = row["CHANNEL_LEVEL_CODE"]
            emp_code = row["EMPLOYEE_CODE"]
            if lvl == "LEVEL03":
                return self.manager_code_mapping.get(emp_code, "")
            if lvl == "LEVEL04":
                return ""
            manager_emp = self._find_manager_emp_within_2_levels(emp_code)
            return self.manager_code_mapping.get(manager_emp, "") if manager_emp else ""

        df["MANAGER_CODE"] = df.apply(assign_manager_code_row, axis=1)
        df["MANAGER_NAME"] = df["MANAGER_CODE"].map(self.manager_code_to_name).fillna("")

    def _assign_team_lead_codes(self, df: pd.DataFrame) -> None:
        team_lead_counter: Dict[str, int] = {}
        level2_df = df[df["CHANNEL_LEVEL_CODE"] == "LEVEL02"]
        for _, row in level2_df.iterrows():
            emp_code = row["EMPLOYEE_CODE"]
            m_code = row["MANAGER_CODE"]
            if m_code:
                team_lead_counter[m_code] = team_lead_counter.get(m_code, 0) + 1
                tl_code = f"{m_code}_TN{team_lead_counter[m_code]:02d}"
                self.team_lead_code_mapping[emp_code] = tl_code
                self.team_lead_code_to_name[tl_code] = row["FULL_NAME"] if pd.notna(row["FULL_NAME"]) else ""
                self.team_lead_to_manager[tl_code] = m_code
                self.team_lead_code_to_emp_code[tl_code] = emp_code

        def assign_team_lead_code_row(row: pd.Series) -> str:
            lvl = row["CHANNEL_LEVEL_CODE"]
            emp_code = row["EMPLOYEE_CODE"]
            if lvl == "LEVEL02":
                return self.team_lead_code_mapping.get(emp_code, "")
            if lvl in ["LEVEL03", "LEVEL04"]:
                return ""
            tl_emp = self._find_team_lead_emp_within_2_levels(emp_code)
            return self.team_lead_code_mapping.get(tl_emp, "") if tl_emp else ""

        df["TEAM_LEAD_CODE"] = df.apply(assign_team_lead_code_row, axis=1)
        df["TEAM_LEAD_NAME"] = df["TEAM_LEAD_CODE"].map(self.team_lead_code_to_name).fillna("")


def build_time_context(config: AppConfig) -> Tuple[TimeContext, datetime, datetime]:
    tz = pytz.timezone(config.timezone)
    now_vn = datetime.now(tz)
    print("=" * 100)
    print("BÁO CÁO PHÂN CẤP BẢO HIỂM - PHIÊN BẢN TỐI ƯU")
    print("=" * 100)
    print(f"\nThời gian hệ thống: {now_vn.strftime('%d/%m/%Y %H:%M:%S')}")

    current_day, current_month, current_year = now_vn.day, now_vn.month, now_vn.year
    days_in_month = calendar.monthrange(current_year, current_month)[1]
    days_worked = current_day - 1

    print(f"Số ngày làm việc (mặc định): {days_worked}/{days_in_month} ngày")

    confirm = input(
        f"\nXác nhận ngày báo cáo {current_day}/{current_month}/{current_year}? (y/n): "
    ).strip().lower()
    if confirm != "y":
        current_day = int(input("Nhập ngày: "))
        current_month = int(input("Nhập tháng: "))
        current_year = int(input("Nhập năm: "))
        days_in_month = calendar.monthrange(current_year, current_month)[1]
        days_worked = current_day - 1

    today_report = datetime(current_year, current_month, current_day)
    report_date_for_premium = today_report - timedelta(days=1)

    print(f"\n✓ Ngày báo cáo: {today_report.strftime('%d/%m/%Y')}")
    print(f"✓ Ngày dùng cho tên file (D-1): {report_date_for_premium.strftime('%d/%m/%Y')}")
    print(f"✓ Số ngày làm việc: {days_worked}/{days_in_month} ngày")

    today = today_report
    yesterday = today - timedelta(days=1)
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    month_for_quarter = report_date_for_premium.month
    quarter_index = (month_for_quarter - 1) // 3 + 1
    quarter_label = f"Q{quarter_index}"
    quarter_months = [3 * (quarter_index - 1) + i for i in range(1, 4)]
    retention_month_labels = [f"Tháng {m}" for m in quarter_months]

    print(f"✓ Thuộc quý: {quarter_label} (các tháng: {', '.join(str(m) for m in quarter_months)})")

    tc = TimeContext(
        today=today,
        yesterday=yesterday,
        week_ago=week_ago,
        month_ago=month_ago,
        days_in_month=days_in_month,
        days_worked=days_worked,
        quarter_index=quarter_index,
        quarter_label=quarter_label,
        quarter_months=quarter_months,
        retention_month_labels=retention_month_labels,
    )
    return tc, today_report, report_date_for_premium


def main() -> None:
    if IS_COLAB:
        drive.mount("/content/drive")
    show_loading("Đang khởi tạo hệ thống...")

    config = AppConfig()
    _ensure_directories(
        [
            config.phi_t_dir,
            config.rawdata_onl_dir,
            config.rawdata_off_dir,
            config.keep_position_dir,
        ]
    )

    time_ctx, today_report, report_date_for_premium = build_time_context(config)
    raw_file = _latest_excel(
        config.rawdata_onl_dir, required=True, description="rawdata online (Cấp onl)"
    )

    if raw_file is None:
        raise FileNotFoundError("Không tìm thấy rawdata online (Cấp onl)")

    processor = InsuranceDataProcessor(
        rawdata_path=raw_file,
        sheet_id=config.sheet_id,
        gid=config.sheet_gid,
        channel_filter="CollaboratorApp",
        excluded_partner=None,
        offline_folder=config.rawdata_off_dir,
    )

    df_premium = processor.process()
    df_premium_all = processor.df_all.copy() if processor.df_all is not None else df_premium.copy()

    premium_output_filename = (
        f"Phí bảo hiểm {report_date_for_premium.day:02d}_"
        f"{report_date_for_premium.month:02d}_"
        f"{report_date_for_premium.year}.xlsx"
    )
    premium_output_path = Path(config.phi_t_dir) / premium_output_filename
    processor.export_to_file(str(premium_output_path))

    quarter_label = time_ctx.quarter_label
    month_for_quarter = report_date_for_premium.month
    keep_q_dir = Path(config.keep_position_dir) / quarter_label
    keep_t_dir = keep_q_dir / f"T{month_for_quarter}"
    _ensure_directories([keep_q_dir, keep_t_dir])

    keep_t_path = keep_t_dir / premium_output_filename
    try:
        shutil.copy2(premium_output_path, keep_t_path)
        print(f"✓ Đã sao chép file phí T vào thư mục giữ hạng: {keep_t_path}")
    except Exception as exc:  # pragma: no cover - log only
        print(f"⚠ Không sao chép được file phí T vào thư mục giữ hạng: {exc}")

    print(f"\n✓ ĐÃ TẠO FILE PHÍ BẢO HIỂM T: {premium_output_filename}")
    # The remainder of the pipeline (hierarchy building and Excel formatting) can be
    # executed by importing and extending this optimized module.

    if IS_COLAB:
        print("\nĐang tải file về máy...")
        files.download(str(premium_output_path))


if __name__ == "__main__":
    main()
